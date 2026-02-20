import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList

wb = openpyxl.Workbook()

# ============ COLORS ============
DARK_BLUE = "0A1628"
BRAND_BLUE = "0066FF"
BRAND_GREEN = "00C9A7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7FAFC"
MEDIUM_GRAY = "E2E8F0"
DARK_TEXT = "333333"
RED = "E53E3E"
GREEN = "38A169"
GOLD = "FFD700"
LIGHT_BLUE_BG = "E8F4FD"
LIGHT_GREEN_BG = "C6F6D5"
LIGHT_RED_BG = "FFE0E0"
PURPLE = "805AD5"
ORANGE = "ED8936"

# ============ STYLES ============
header_font = Font(name='Calibri', bold=True, size=11, color=WHITE)
title_font = Font(name='Calibri', bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
normal_font = Font(name='Calibri', size=11, color=DARK_TEXT)
bold_font = Font(name='Calibri', bold=True, size=11, color=DARK_TEXT)
small_font = Font(name='Calibri', size=10, color="666666")
number_font = Font(name='Calibri', size=11, color=DARK_TEXT)
big_green = Font(name='Calibri', bold=True, size=13, color=GREEN)
big_blue = Font(name='Calibri', bold=True, size=13, color=BRAND_BLUE)
big_red = Font(name='Calibri', bold=True, size=13, color=RED)
milestone_font = Font(name='Calibri', bold=True, size=11, color=PURPLE)

dark_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
blue_fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type='solid')
green_fill = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE_BG, end_color=LIGHT_BLUE_BG, fill_type='solid')
light_green_fill = PatternFill(start_color=LIGHT_GREEN_BG, end_color=LIGHT_GREEN_BG, fill_type='solid')
light_red_fill = PatternFill(start_color=LIGHT_RED_BG, end_color=LIGHT_RED_BG, fill_type='solid')
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
gold_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type='solid')
purple_fill = PatternFill(start_color="E9D8FD", end_color="E9D8FD", fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border


def style_cell(ws, row, col, val, font=normal_font, align=center, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# =====================================================
# SHEET 1: 12-MONTH TRAFFIC PROJECTIONS
# =====================================================
ws1 = wb.active
ws1.title = "Traffic Projections"
ws1.sheet_properties.tabColor = BRAND_BLUE

col_widths = [5, 14, 18, 16, 14, 16, 18, 16, 14, 30]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells('A1:J1')
ws1.cell(row=1, column=1, value="OJCOMMERCE — 12-MONTH REFERRAL TRAFFIC PROJECTION").font = title_font
ws1.cell(row=1, column=1).fill = light_blue_fill
ws1.cell(row=1, column=1).alignment = center

ws1.merge_cells('A2:J2')
ws1.cell(row=2, column=1, value="Website: ojcommerce.com | Current Organic Users: 122,173/month | Current Referral: 0.28% (~342/month) | Target: 5-10%").font = small_font
ws1.cell(row=2, column=1).alignment = center

# Headers
headers = ["#", "Month", "New Backlinks\nAcquired", "Cumulative\nBacklinks", "Avg DR of\nNew Links",
           "Est. Referral\nVisitors/Month", "Referral Traffic\n% of Total", "Growth vs\nBaseline", "Phase", "Milestones & Notes"]

style_header(ws1, 4, 1, 10)
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)

# 12-month data
# (month_num, month_name, new_links, cumulative, avg_dr, ref_visitors, ref_pct, growth, phase, notes)
data = [
    (0, "Current State", 0, 0, "-", 342, "0.28%", "-", "BASELINE", "Starting point: 0.28% referral traffic. No existing backlinks."),
    (1, "Month 1", 4, 4, "62", 520, "0.43%", "+52%", "FOUNDATION", "Audit complete. First 4 DR 60+ links live. Twice-weekly reporting begins."),
    (2, "Month 2", 6, 10, "64", 890, "0.73%", "+160%", "FOUNDATION", "Outreach scaled. 6 new placements. First measurable referral traffic growth."),
    (3, "Month 3", 8, 18, "65", 1500, "1.23%", "+339%", "SCALE", "Pipeline established. Content placements generating consistent referral visits."),
    (4, "Month 4", 10, 28, "66", 2400, "1.96%", "+602%", "SCALE", "Momentum building. Higher DR placements secured. Brand visibility improving."),
    (5, "Month 5", 10, 38, "67", 3500, "2.87%", "+923%", "SCALE", "Referral traffic compounding. Category page links driving product traffic."),
    (6, "Month 6", 12, 50, "68", 4800, "3.93%", "+1303%", "OPTIMIZE", "6-month review point. Strategy refinement based on top-performing channels."),
    (7, "Month 7", 12, 62, "69", 6100, "4.99%", "+1684%", "OPTIMIZE", "MINIMUM TARGET HIT: 5% referral traffic achieved."),
    (8, "Month 8", 12, 74, "70", 7400, "6.06%", "+2064%", "OPTIMIZE", "Exceeding minimum target. High-DR placements driving significant traffic."),
    (9, "Month 9", 14, 88, "71", 8800, "7.20%", "+2473%", "ACCELERATE", "Premium DR 70+ links accelerating growth. Digital PR results landing."),
    (10, "Month 10", 14, 102, "71", 10200, "8.35%", "+2882%", "ACCELERATE", "Strong referral traffic. Brand visibility keywords ranking on page 1."),
    (11, "Month 11", 14, 116, "72", 11400, "9.33%", "+3233%", "ACCELERATE", "Approaching stretch target. Compounding effect of 100+ quality backlinks."),
    (12, "Month 12", 15, 131, "72", 12200, "9.99%", "+3467%", "TARGET HIT", "STRETCH TARGET HIT: ~10% referral traffic. 131 DR 60+ backlinks total."),
]

for i, (num, month, new_links, cumul, avg_dr, ref_visitors, ref_pct, growth, phase, notes) in enumerate(data):
    row = 5 + i
    alt_fill = light_gray_fill if i % 2 == 0 else white_fill

    # Highlight special rows
    if phase == "BASELINE":
        row_fill = light_red_fill
    elif "TARGET HIT" in phase or "MINIMUM TARGET" in notes:
        row_fill = light_green_fill
    elif phase == "ACCELERATE":
        row_fill = purple_fill
    else:
        row_fill = alt_fill

    style_cell(ws1, row, 1, num if num > 0 else "-", bold_font, center, row_fill)
    style_cell(ws1, row, 2, month, bold_font, center, row_fill)
    style_cell(ws1, row, 3, new_links, number_font, center, row_fill)
    style_cell(ws1, row, 4, cumul, big_blue, center, row_fill)
    style_cell(ws1, row, 5, avg_dr, number_font, center, row_fill)
    style_cell(ws1, row, 6, f"{ref_visitors:,}", big_green if ref_visitors >= 6100 else bold_font, center, row_fill)
    style_cell(ws1, row, 7, ref_pct, big_green if ref_visitors >= 6100 else bold_font, center, row_fill)
    style_cell(ws1, row, 8, growth, normal_font, center, row_fill)
    style_cell(ws1, row, 9, phase, milestone_font if "TARGET" in phase or "ACCELERATE" in phase else normal_font, center, row_fill)
    style_cell(ws1, row, 10, notes, small_font, left, row_fill)

# Summary row
row = 18
ws1.merge_cells(f'A{row}:B{row}')
style_cell(ws1, row, 1, "12-MONTH TOTAL", Font(name='Calibri', bold=True, size=12, color=WHITE), center, dark_fill)
ws1.cell(row=row, column=2).fill = dark_fill
ws1.cell(row=row, column=2).border = thin_border
style_cell(ws1, row, 3, 131, Font(name='Calibri', bold=True, size=12, color=WHITE), center, dark_fill)
style_cell(ws1, row, 4, 131, Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws1, row, 5, "Avg 68", Font(name='Calibri', bold=True, size=11, color=WHITE), center, dark_fill)
style_cell(ws1, row, 6, "12,200", Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws1, row, 7, "~10%", Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws1, row, 8, "+3,467%", Font(name='Calibri', bold=True, size=11, color=BRAND_GREEN), center, dark_fill)
style_cell(ws1, row, 9, "COMPLETE", Font(name='Calibri', bold=True, size=11, color=BRAND_GREEN), center, dark_fill)
style_cell(ws1, row, 10, "Both minimum (5%) and stretch (10%) targets achieved", Font(name='Calibri', bold=True, size=10, color=WHITE), left, dark_fill)

# Key assumptions
row = 20
ws1.merge_cells(f'A{row}:J{row}')
ws1.cell(row=row, column=1, value="KEY ASSUMPTIONS").font = subtitle_font
ws1.cell(row=row, column=1).fill = light_blue_fill

assumptions = [
    "Organic traffic baseline remains stable at ~122,173 users/month (conservative — likely to grow with improved SEO)",
    "All backlinks are DR 60+ from furniture/home niche sites with real US traffic (1,000+ daily users)",
    "40% of links target homepage/brand, 60% target category/product/blog pages (as Alkesh specified)",
    "Referral traffic per link increases over time as pages age and accumulate more authority",
    "Projections are conservative — actual results may exceed these based on placement quality",
    "GA4 referral traffic reports to be shared by client once project starts for precise tracking",
    "Month 7 is the earliest realistic point to hit 5% minimum target; Month 12 for 10% stretch target",
]

for i, assumption in enumerate(assumptions):
    row = 21 + i
    ws1.merge_cells(f'A{row}:J{row}')
    ws1.cell(row=row, column=1, value=f"  {i+1}. {assumption}").font = small_font


# =====================================================
# SHEET 2: LINK DISTRIBUTION PLAN
# =====================================================
ws2 = wb.create_sheet("Link Distribution")
ws2.sheet_properties.tabColor = BRAND_GREEN

col_widths2 = [30, 15, 15, 20, 35]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value="OJCOMMERCE — LINK DISTRIBUTION PLAN").font = title_font
ws2.cell(row=1, column=1).fill = light_blue_fill
ws2.cell(row=1, column=1).alignment = center

ws2.merge_cells('A2:E2')
ws2.cell(row=2, column=1, value="Alkesh specified: 40% focused on brand + homepage authority").font = Font(name='Calibri', bold=True, size=11, color=GREEN)
ws2.cell(row=2, column=1).alignment = center

# Target page distribution
style_header(ws2, 4, 1, 5)
dist_headers = ["Target Page Type", "Link Share %", "Links in 12 Months\n(of 131 total)", "Sample Target URLs", "Anchor Text Examples"]
for i, h in enumerate(dist_headers, 1):
    ws2.cell(row=4, column=i, value=h)

dist_data = [
    ("Homepage (brand authority)", "40%", 52, "ojcommerce.com", '"OJCommerce", "OJ Commerce", "ojcommerce.com", "visit OJCommerce"'),
    ("Category Pages", "30%", 39, "ojcommerce.com/bedroom-furniture\nojcommerce.com/living-room\nojcommerce.com/dining-furniture", '"bedroom furniture", "OJCommerce living room sets", "affordable dining furniture"'),
    ("Product Pages", "20%", 26, "ojcommerce.com/product/[specific]\n(To be shared by Alkesh)", '"buy [product name] online", "OJ Commerce [product]", "[product] review"'),
    ("Blog / Content Pages", "10%", 14, "ojcommerce.com/blog/[articles]\n(To be created as part of strategy)", '"furniture buying guide", "home design tips", "learn more"'),
]

for i, (page, share, count, urls, anchors) in enumerate(dist_data):
    row = 5 + i
    fill = light_blue_fill if i == 0 else (light_gray_fill if i % 2 == 1 else white_fill)
    style_cell(ws2, row, 1, page, bold_font, left, fill)
    style_cell(ws2, row, 2, share, Font(name='Calibri', bold=True, size=14, color=BRAND_BLUE), center, fill)
    style_cell(ws2, row, 3, count, big_blue, center, fill)
    style_cell(ws2, row, 4, urls, small_font, left, fill)
    style_cell(ws2, row, 5, anchors, small_font, left, fill)

# Total row
row = 9
style_cell(ws2, row, 1, "TOTAL", Font(name='Calibri', bold=True, size=12, color=WHITE), center, dark_fill)
style_cell(ws2, row, 2, "100%", Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws2, row, 3, 131, Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws2, row, 4, "", header_font, center, dark_fill)
style_cell(ws2, row, 5, "", header_font, center, dark_fill)

# Anchor text distribution
row = 11
ws2.merge_cells(f'A{row}:E{row}')
ws2.cell(row=row, column=1, value="ANCHOR TEXT DISTRIBUTION").font = subtitle_font
ws2.cell(row=row, column=1).fill = light_blue_fill

style_header(ws2, 12, 1, 5)
anchor_headers = ["Anchor Type", "Target %", "Links in 12 Months", "Examples", "Why This Ratio"]
for i, h in enumerate(anchor_headers, 1):
    ws2.cell(row=12, column=i, value=h)

anchor_data = [
    ("Branded", "35-40%", "46-52", '"OJCommerce", "OJ Commerce"', "Builds brand authority — Alkesh's primary goal"),
    ("Branded + Keyword", "15-20%", "20-26", '"OJCommerce furniture", "OJ Commerce bedroom sets"', "Combines brand with target keywords naturally"),
    ("Exact Match Keyword", "5-10%", "7-13", '"buy dining table online", "furniture store USA"', "Targets high-intent terms — keep low to avoid penalties"),
    ("Partial Match", "10-15%", "13-20", '"affordable bedroom furniture", "quality home furniture"', "Natural variation signals editorial placement"),
    ("Generic", "5-10%", "7-13", '"click here", "learn more", "visit website"', "Natural pattern in organic link profiles"),
    ("Naked URL", "10-15%", "13-20", '"ojcommerce.com", "www.ojcommerce.com"', "Common in editorial mentions and citations"),
]

for i, (atype, pct, count, examples, why) in enumerate(anchor_data):
    row = 13 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws2, row, 1, atype, bold_font, left, fill)
    style_cell(ws2, row, 2, pct, big_blue, center, fill)
    style_cell(ws2, row, 3, count, bold_font, center, fill)
    style_cell(ws2, row, 4, examples, small_font, left, fill)
    style_cell(ws2, row, 5, why, small_font, left, fill)


# =====================================================
# SHEET 3: MONTHLY COST PROJECTIONS
# =====================================================
ws3 = wb.create_sheet("Cost Projections")
ws3.sheet_properties.tabColor = ORANGE

col_widths3 = [14, 14, 14, 14, 16, 18, 18, 30]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells('A1:H1')
ws3.cell(row=1, column=1, value="OJCOMMERCE — MONTHLY COST PROJECTION (PER-LINK MODEL)").font = title_font
ws3.cell(row=1, column=1).fill = light_blue_fill
ws3.cell(row=1, column=1).alignment = center

ws3.merge_cells('A2:H2')
ws3.cell(row=2, column=1, value="Fill in YOUR per-link prices in the yellow cells below. All other values auto-calculate.").font = Font(name='Calibri', bold=True, size=11, color=RED)
ws3.cell(row=2, column=1).alignment = center

# Price inputs
row = 4
ws3.merge_cells(f'A{row}:B{row}')
ws3.cell(row=row, column=1, value="YOUR PER-LINK PRICING:").font = subtitle_font

style_header(ws3, 5, 1, 3)
ws3.cell(row=5, column=1, value="DR Tier")
ws3.cell(row=5, column=2, value="Your Price")
ws3.cell(row=5, column=3, value="Market Range")

tiers = [
    ("DR 60-69", "", "$250-$400"),
    ("DR 70-79", "", "$400-$600"),
    ("DR 80+", "", "$600-$1,000+"),
]

for i, (tier, price, market) in enumerate(tiers):
    row = 6 + i
    style_cell(ws3, row, 1, tier, bold_font, center, light_gray_fill)
    cell = style_cell(ws3, row, 2, price, Font(name='Calibri', bold=True, size=14, color=RED), center, gold_fill)
    cell.number_format = '$#,##0'
    style_cell(ws3, row, 3, market, small_font, center, light_gray_fill)

ws3.merge_cells('D4:H4')
ws3.cell(row=4, column=4, value="VOLUME DISCOUNT STRUCTURE:").font = subtitle_font

style_header(ws3, 5, 4, 6)
ws3.cell(row=5, column=4, value="Volume")
ws3.cell(row=5, column=5, value="Discount")
ws3.cell(row=5, column=6, value="Incentive")

vol_data = [
    ("1-9 links/month", "Standard pricing", "Base rate"),
    ("10-14 links/month", "10% discount", "Encourages consistent volume"),
    ("15+ links/month", "15% discount", "Maximum savings for commitment"),
]

for i, (vol, disc, inc) in enumerate(vol_data):
    row = 6 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws3, row, 4, vol, bold_font, center, fill)
    style_cell(ws3, row, 5, disc, Font(name='Calibri', bold=True, size=11, color=GREEN), center, fill)
    style_cell(ws3, row, 6, inc, small_font, left, fill)

# Monthly breakdown
row = 10
ws3.merge_cells(f'A{row}:H{row}')
ws3.cell(row=row, column=1, value="MONTHLY COST BREAKDOWN (Fill in prices above — then calculate totals below)").font = subtitle_font
ws3.cell(row=row, column=1).fill = light_blue_fill

style_header(ws3, 11, 1, 8)
cost_headers = ["Month", "DR 60-69\nLinks", "DR 70-79\nLinks", "DR 80+\nLinks", "Total\nLinks", "Est. Cost\n(fill in)", "Cumulative\nCost", "Notes"]
for i, h in enumerate(cost_headers, 1):
    ws3.cell(row=11, column=i, value=h)

monthly_links = [
    ("Month 1", 3, 1, 0, 4, "", "", "Foundation phase — audit + first placements"),
    ("Month 2", 4, 2, 0, 6, "", "", "Scaling outreach — pipeline established"),
    ("Month 3", 5, 2, 1, 8, "", "", "First premium DR 80+ link acquired"),
    ("Month 4", 6, 3, 1, 10, "", "", "Full velocity — consistent output"),
    ("Month 5", 5, 3, 2, 10, "", "", "Higher DR mix — quality increasing"),
    ("Month 6", 6, 4, 2, 12, "", "", "6-month review — optimize strategy"),
    ("Month 7", 5, 4, 3, 12, "", "", "MINIMUM TARGET (5%) hit this month"),
    ("Month 8", 5, 4, 3, 12, "", "", "Exceeding target — premium links accelerating"),
    ("Month 9", 5, 5, 4, 14, "", "", "Highest quality month — strong DR 80+ pipeline"),
    ("Month 10", 5, 5, 4, 14, "", "", "Brand keywords ranking improvements visible"),
    ("Month 11", 5, 5, 4, 14, "", "", "Approaching 10% stretch target"),
    ("Month 12", 6, 5, 4, 15, "", "", "STRETCH TARGET (~10%) achieved"),
]

for i, (month, dr60, dr70, dr80, total, cost, cumul, notes) in enumerate(monthly_links):
    row = 12 + i
    fill = light_green_fill if "TARGET" in notes else (light_gray_fill if i % 2 == 0 else white_fill)
    style_cell(ws3, row, 1, month, bold_font, center, fill)
    style_cell(ws3, row, 2, dr60, number_font, center, fill)
    style_cell(ws3, row, 3, dr70, number_font, center, fill)
    style_cell(ws3, row, 4, dr80, number_font, center, fill)
    style_cell(ws3, row, 5, total, big_blue, center, fill)
    style_cell(ws3, row, 6, cost, normal_font, center, gold_fill)  # yellow for manual entry
    style_cell(ws3, row, 7, cumul, normal_font, center, gold_fill)
    style_cell(ws3, row, 8, notes, small_font, left, fill)

# Total row
row = 24
style_cell(ws3, row, 1, "12-MONTH TOTAL", Font(name='Calibri', bold=True, size=12, color=WHITE), center, dark_fill)
style_cell(ws3, row, 2, 60, Font(name='Calibri', bold=True, size=11, color=WHITE), center, dark_fill)
style_cell(ws3, row, 3, 43, Font(name='Calibri', bold=True, size=11, color=WHITE), center, dark_fill)
style_cell(ws3, row, 4, 28, Font(name='Calibri', bold=True, size=11, color=WHITE), center, dark_fill)
style_cell(ws3, row, 5, 131, Font(name='Calibri', bold=True, size=12, color=BRAND_GREEN), center, dark_fill)
style_cell(ws3, row, 6, "Calculate", Font(name='Calibri', bold=True, size=11, color=GOLD), center, dark_fill)
style_cell(ws3, row, 7, "Calculate", Font(name='Calibri', bold=True, size=11, color=GOLD), center, dark_fill)
style_cell(ws3, row, 8, "Total investment for 131 DR 60+ backlinks over 12 months", Font(name='Calibri', size=10, color=WHITE), left, dark_fill)

# DR mix summary
row = 26
ws3.merge_cells(f'A{row}:H{row}')
ws3.cell(row=row, column=1, value="DR MIX SUMMARY: 60 links (DR 60-69) = 46% | 43 links (DR 70-79) = 33% | 28 links (DR 80+) = 21% — Quality increases over time as relationships mature").font = small_font
ws3.cell(row=row, column=1).fill = light_blue_fill


# =====================================================
# SHEET 4: ROI CALCULATOR
# =====================================================
ws4 = wb.create_sheet("ROI Calculator")
ws4.sheet_properties.tabColor = GREEN

col_widths4 = [35, 25, 30]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells('A1:C1')
ws4.cell(row=1, column=1, value="OJCOMMERCE — ROI & VALUE CALCULATOR").font = title_font
ws4.cell(row=1, column=1).fill = light_blue_fill
ws4.cell(row=1, column=1).alignment = center

ws4.merge_cells('A2:C2')
ws4.cell(row=2, column=1, value="Show Alkesh the VALUE of referral traffic — not just the cost of links").font = Font(name='Calibri', bold=True, size=11, color=GREEN)
ws4.cell(row=2, column=1).alignment = center

# Current state
row = 4
ws4.merge_cells(f'A{row}:C{row}')
ws4.cell(row=row, column=1, value="CURRENT STATE").font = subtitle_font
ws4.cell(row=row, column=1).fill = light_red_fill

metrics = [
    ("Monthly organic users", "122,173", "From Alkesh's checklist"),
    ("Current referral traffic %", "0.28%", "Nearly zero"),
    ("Current referral visitors/month", "~342", "122,173 x 0.28%"),
    ("Current referral revenue contribution", "Minimal", "Almost no revenue from referrals"),
]

style_header(ws4, 5, 1, 3)
ws4.cell(row=5, column=1, value="Metric")
ws4.cell(row=5, column=2, value="Current Value")
ws4.cell(row=5, column=3, value="Source")

for i, (metric, val, source) in enumerate(metrics):
    row = 6 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws4, row, 1, metric, bold_font, left, fill)
    style_cell(ws4, row, 2, val, big_red, center, fill)
    style_cell(ws4, row, 3, source, small_font, left, fill)

# Projected state
row = 11
ws4.merge_cells(f'A{row}:C{row}')
ws4.cell(row=row, column=1, value="PROJECTED STATE (After 12 Months)").font = subtitle_font
ws4.cell(row=row, column=1).fill = light_green_fill

projected = [
    ("Monthly referral visitors (5% target)", "~6,100/month", "Minimum target"),
    ("Monthly referral visitors (10% target)", "~12,200/month", "Stretch target"),
    ("Annual referral visitors (5%)", "~73,200/year", "New visitors from backlinks"),
    ("Annual referral visitors (10%)", "~146,400/year", "New visitors from backlinks"),
    ("Total DR 60+ backlinks acquired", "131", "Over 12 months"),
    ("Average domain rating of links", "68", "High authority, niche relevant"),
    ("Estimated DR improvement for OJCommerce", "+8-12 points", "Significant authority boost"),
]

style_header(ws4, 12, 1, 3)
ws4.cell(row=12, column=1, value="Metric")
ws4.cell(row=12, column=2, value="Projected Value")
ws4.cell(row=12, column=3, value="Notes")

for i, (metric, val, notes) in enumerate(projected):
    row = 13 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws4, row, 1, metric, bold_font, left, fill)
    style_cell(ws4, row, 2, val, big_green, center, fill)
    style_cell(ws4, row, 3, notes, small_font, left, fill)

# Value calculation
row = 21
ws4.merge_cells(f'A{row}:C{row}')
ws4.cell(row=row, column=1, value="VALUE CALCULATION — What is this referral traffic worth?").font = subtitle_font
ws4.cell(row=row, column=1).fill = gold_fill

style_header(ws4, 22, 1, 3)
ws4.cell(row=22, column=1, value="Scenario")
ws4.cell(row=22, column=2, value="Calculation")
ws4.cell(row=22, column=3, value="Annual Value")

value_data = [
    ("If OJCommerce conversion rate = 1%", "73,200 visitors x 1% = 732 orders/year", "Fill in: 732 x avg order value"),
    ("If OJCommerce conversion rate = 2%", "73,200 visitors x 2% = 1,464 orders/year", "Fill in: 1,464 x avg order value"),
    ("If avg order value = $500", "732 orders x $500 = $366,000/year (at 1% conv)", "$366,000 annual revenue from referrals"),
    ("If avg order value = $500", "1,464 orders x $500 = $732,000/year (at 2% conv)", "$732,000 annual revenue from referrals"),
    ("CPC equivalent (what this traffic would cost in ads)", "73,200 visitors x ~$1.50 CPC", "$109,800/year saved vs paid ads"),
]

for i, (scenario, calc, value) in enumerate(value_data):
    row = 23 + i
    fill = light_green_fill if i >= 2 else (light_gray_fill if i % 2 == 0 else white_fill)
    style_cell(ws4, row, 1, scenario, bold_font, left, fill)
    style_cell(ws4, row, 2, calc, normal_font, left, fill)
    style_cell(ws4, row, 3, value, big_green if i >= 2 else normal_font, left, fill)

row = 29
ws4.merge_cells(f'A{row}:C{row}')
ws4.cell(row=row, column=1, value="KEY TALKING POINT: \"Even at conservative estimates, the referral traffic from 131 DR 60+ backlinks could generate $366K+ in annual revenue — far exceeding the link building investment.\"").font = Font(name='Calibri', bold=True, size=11, color=GREEN)
ws4.cell(row=row, column=1).fill = light_green_fill
ws4.cell(row=row, column=1).alignment = left


# =====================================================
# SHEET 5: DEMO CHEAT SHEET
# =====================================================
ws5 = wb.create_sheet("Demo Cheat Sheet")
ws5.sheet_properties.tabColor = RED

col_widths5 = [12, 30, 50]
for i, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value="FRIDAY DEMO — MINUTE-BY-MINUTE CHEAT SHEET").font = title_font
ws5.cell(row=1, column=1).fill = light_red_fill
ws5.cell(row=1, column=1).alignment = center

ws5.merge_cells('A2:C2')
ws5.cell(row=2, column=1, value="Keep this open on a second screen during the demo. 40 minutes total.").font = Font(name='Calibri', bold=True, size=11, color=RED)
ws5.cell(row=2, column=1).alignment = center

style_header(ws5, 4, 1, 3)
ws5.cell(row=4, column=1, value="Time")
ws5.cell(row=4, column=2, value="Section")
ws5.cell(row=4, column=3, value="What to Say / Show")

cheat = [
    ("0:00-3:00", "OPENING — His Data", '"Alkesh, your referral traffic is 0.28% — that\'s 342 sessions/month. Your target is 5-10%, meaning 6,100-12,200 sessions. Let me show you exactly how we get there."\n\nShow: Current State slide with 122,173 users and 0.28% referral'),
    ("3:00-12:00", "TOOL DEMO", "Show how you:\n1. Source DR 60+ prospects in furniture niche\n2. Verify real US traffic (not bots)\n3. Filter out risky/irrelevant domains\n4. Sample verification report for one domain\n\nKEY LINE: \"This is exactly how we prevent what happened with your previous agency. Every domain goes through this verification.\""),
    ("12:00-20:00", "COMPETITOR ANALYSIS", "Show Ahrefs data for:\n- One Stop Bedroom (direct competitor — focus here)\n- Rooms To Go (DR 66, 158K backlinks, 6.6K domains)\n- Bed Bath & Beyond\n\nShow BACKLINK GAP: Sites linking to them but NOT OJCommerce\n\"These are our Month 1 targets — sites already linking to your competitors.\""),
    ("20:00-27:00", "OJCOMMERCE STRATEGY", "Show Link Distribution: 40% homepage / 30% category / 20% product / 10% blog\n\nShow Traffic Projections spreadsheet:\n- Month 1: 4 links, 0.43% referral\n- Month 3: 18 cumulative, 1.23%\n- Month 7: 5% TARGET HIT\n- Month 12: ~10% STRETCH TARGET\n\nKEY LINE: \"These are conservative projections.\""),
    ("27:00-33:00", "CASE STUDIES", "Show 2-3 examples. If no furniture cases, use eCommerce:\n\"Similar competitive dynamics to OJCommerce\"\n\nHighlight: referral traffic growth, DR improvement, ranking gains"),
    ("33:00-38:00", "PRICING & CLOSE", "Show 3-tier per-link pricing (DR 60-70, 70-80, 80+)\nEmphasize:\n- Per-link = he only pays for quality\n- No payment until link is live\n- He approves EVERY placement\n- Volume discount for 10+/month\n\nCLOSE: \"I'd recommend 8-10 links/month across the 3 tiers. Would you like to move forward?\""),
    ("38:00-40:00", "NEXT STEPS", "If YES:\n1. Send agreement within 24 hours\n2. He shares GA4 reports + priority URLs\n3. Audit + prospecting begins Week 1\n4. First outreach Week 2\n5. First twice-weekly update lands next week\n\nIf NEED APPROVAL: \"Would it help to include the business owner in a 10-min follow-up call?\""),
]

for i, (time, section, script) in enumerate(cheat):
    row = 5 + i
    if "CLOSE" in section:
        fill = light_green_fill
    elif "OPENING" in section:
        fill = light_blue_fill
    else:
        fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws5, row, 1, time, bold_font, center, fill)
    style_cell(ws5, row, 2, section, Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE), center, fill)
    cell = style_cell(ws5, row, 3, script, normal_font, left, fill)
    ws5.row_dimensions[row].height = 100

# Quick stats
row = 13
ws5.merge_cells(f'A{row}:C{row}')
ws5.cell(row=row, column=1, value="QUICK STATS TO REMEMBER DURING THE DEMO").font = subtitle_font
ws5.cell(row=row, column=1).fill = gold_fill

stats = [
    ("Current referral traffic", "0.28% = ~342 visitors/month"),
    ("Target", "5-10% = 6,100-12,200 visitors/month"),
    ("Growth needed", "17x (minimum) to 35x (stretch)"),
    ("DR requirement", "60+ only (strict)"),
    ("Homepage link allocation", "40% (Alkesh specified)"),
    ("Brand names for anchors", '"OJCommerce" and "OJ Commerce"'),
    ("Competitors (top 3)", "One Stop Bedroom, Rooms To Go, Bed Bath & Beyond"),
    ("Pricing model", "Per-link (not retainer)"),
    ("Reporting", "Twice weekly (Tuesday + Friday)"),
    ("Start date", "Immediate"),
    ("Decision maker", "Business Owner"),
    ("Past frustration", "Volume over quality, no measurable referral traffic"),
    ("Avoid niches", "Political, adult, gambling, unrelated"),
    ("Rooms To Go stats", "DR 66 | 158K backlinks | 6.6K referring domains"),
]

style_header(ws5, 14, 1, 2)
ws5.cell(row=14, column=1, value="Stat")
ws5.cell(row=14, column=2, value="Value")

for i, (stat, val) in enumerate(stats):
    row = 15 + i
    fill = light_gray_fill if i % 2 == 0 else white_fill
    style_cell(ws5, row, 1, stat, bold_font, left, fill)
    style_cell(ws5, row, 2, val, Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE), left, fill)


# ============ SAVE ============
filepath = "/Users/shamique/Downloads/shridhar/Elevro/OJCommerce_Traffic_Projections.xlsx"
wb.save(filepath)
print(f"File saved: {filepath}")
