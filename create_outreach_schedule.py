import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ COLORS ============
DARK_BLUE = "0A1628"
BRAND_BLUE = "0066FF"
BRAND_GREEN = "00C9A7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7FAFC"
MEDIUM_GRAY = "E2E8F0"
DARK_TEXT = "333333"
RED = "E53E3E"
GREEN = "38A169"
LIGHT_BLUE_BG = "E8F4FD"
LIGHT_GREEN_BG = "C6F6D5"
LIGHT_RED_BG = "FFE0E0"
PURPLE = "805AD5"
ORANGE = "ED8936"
YELLOW_BG = "FFFBEB"
PINK_BG = "FFF0F5"
MONDAY_BG = "E8F4FD"
TUESDAY_BG = "E9D8FD"
WEDNESDAY_BG = "FED7AA"
THURSDAY_BG = "C6F6D5"
FRIDAY_BG = "FEEBC8"

# ============ STYLES ============
header_font = Font(name='Calibri', bold=True, size=11, color=WHITE)
title_font = Font(name='Calibri', bold=True, size=16, color=DARK_BLUE)
subtitle_font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
normal_font = Font(name='Calibri', size=11, color=DARK_TEXT)
bold_font = Font(name='Calibri', bold=True, size=11, color=DARK_TEXT)
small_font = Font(name='Calibri', size=10, color="666666")
big_blue = Font(name='Calibri', bold=True, size=13, color=BRAND_BLUE)
big_green = Font(name='Calibri', bold=True, size=13, color=GREEN)
big_red = Font(name='Calibri', bold=True, size=13, color=RED)

dark_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
blue_fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE_BG, end_color=LIGHT_BLUE_BG, fill_type='solid')
light_green_fill = PatternFill(start_color=LIGHT_GREEN_BG, end_color=LIGHT_GREEN_BG, fill_type='solid')
light_red_fill = PatternFill(start_color=LIGHT_RED_BG, end_color=LIGHT_RED_BG, fill_type='solid')
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')
yellow_fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type='solid')
pink_fill = PatternFill(start_color=PINK_BG, end_color=PINK_BG, fill_type='solid')
mon_fill = PatternFill(start_color=MONDAY_BG, end_color=MONDAY_BG, fill_type='solid')
tue_fill = PatternFill(start_color=TUESDAY_BG, end_color=TUESDAY_BG, fill_type='solid')
wed_fill = PatternFill(start_color=WEDNESDAY_BG, end_color=WEDNESDAY_BG, fill_type='solid')
thu_fill = PatternFill(start_color=THURSDAY_BG, end_color=THURSDAY_BG, fill_type='solid')
fri_fill = PatternFill(start_color=FRIDAY_BG, end_color=FRIDAY_BG, fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border


def sc(ws, row, col, val, font=normal_font, align=center, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell


# =====================================================
# SHEET 1: DAILY OUTREACH SCHEDULE - MONTH 1
# =====================================================
ws1 = wb.active
ws1.title = "Month 1 Daily Schedule"
ws1.sheet_properties.tabColor = BRAND_BLUE

widths = [6, 12, 12, 10, 22, 30, 14, 14, 30]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells('A1:I1')
ws1.cell(row=1, column=1, value="OJCOMMERCE — MONTH 1 DAILY OUTREACH SCHEDULE").font = title_font
ws1.cell(row=1, column=1).fill = light_blue_fill
ws1.cell(row=1, column=1).alignment = center

ws1.merge_cells('A2:I2')
ws1.cell(row=2, column=1, value="Target: 4 backlinks | ~80 outreach emails | 4 emails/day (Mon-Fri) | Weeks 1-2: Setup & Research | Weeks 3-4: Active Outreach").font = small_font
ws1.cell(row=2, column=1).alignment = center

# Summary boxes
ws1.merge_cells('A3:C3')
sc(ws1, 3, 1, "Month 1 Target: 4 Links", Font(name='Calibri', bold=True, size=12, color=WHITE), center, blue_fill)
ws1.cell(row=3, column=2).fill = blue_fill
ws1.cell(row=3, column=3).fill = blue_fill
ws1.merge_cells('D3:F3')
sc(ws1, 3, 4, "Emails Required: ~80", Font(name='Calibri', bold=True, size=12, color=WHITE), center, PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid'))
ws1.cell(row=3, column=5).fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
ws1.cell(row=3, column=6).fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
ws1.merge_cells('G3:I3')
sc(ws1, 3, 7, "Daily Rate: 4 emails/day", Font(name='Calibri', bold=True, size=12, color=WHITE), center, PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid'))
ws1.cell(row=3, column=8).fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid')
ws1.cell(row=3, column=9).fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type='solid')

# Headers
style_header(ws1, 5, 1, 9)
headers = ["Day", "Date", "Day of Week", "Time Block", "Task Type", "Specific Actions", "Emails to Send", "Follow-ups", "Notes / Tips"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=5, column=i, value=h)

# Day fills by weekday
day_fills = {
    "Monday": mon_fill,
    "Tuesday": tue_fill,
    "Wednesday": wed_fill,
    "Thursday": thu_fill,
    "Friday": fri_fill,
    "WEEKEND": light_gray_fill,
}

# WEEK 1 - Setup & Research
week1_data = [
    # Week 1 header
    ("", "", "WEEK 1: SETUP, RESEARCH & AUDIT", "", "", "", "", "", ""),
    ("1", "Day 1", "Monday", "9AM-12PM\n1PM-3PM", "SETUP & AUDIT",
     "1. Run OJCommerce through Ahrefs\n2. Document current DR, backlinks, referring domains\n3. Set up outreach tracking spreadsheet\n4. Set up email templates",
     "0", "0", "No emails yet — foundation day. Get all tools and templates ready."),
    ("2", "Day 2", "Tuesday", "9AM-12PM\n1PM-3PM", "COMPETITOR RESEARCH",
     "1. Run One Stop Bedroom through Ahrefs\n2. Run Rooms To Go through Ahrefs\n3. Run Bed Bath & Beyond through Ahrefs\n4. Document all competitor data",
     "0", "0", "REPORT TO ALKESH: Tuesday update #1 — share audit progress."),
    ("3", "Day 3", "Wednesday", "9AM-12PM\n1PM-3PM", "COMPETITOR RESEARCH",
     "1. Run Wayfair through Ahrefs\n2. Run Ashley Furniture through Ahrefs\n3. Run Backlink Gap analysis (all 5 vs OJCommerce)\n4. Identify top 20 gap opportunities",
     "0", "0", "Focus on finding DR 60+ sites that link to competitors but NOT OJCommerce."),
    ("4", "Day 4", "Thursday", "9AM-12PM\n1PM-3PM", "PROSPECT LIST BUILDING",
     "1. Research 25 potential target websites\n2. Verify each: DR 60+, real US traffic, furniture niche\n3. Find contact emails using Hunter.io\n4. Add to prospect spreadsheet",
     "0", "0", "Quality over quantity. Only add sites that meet ALL criteria."),
    ("5", "Day 5", "Friday", "9AM-12PM\n1PM-3PM", "PROSPECT LIST + FIRST EMAILS",
     "1. Research 15 more prospect websites\n2. Finalize Week 1 prospect list (40 prospects)\n3. Send FIRST 4 outreach emails to best prospects\n4. Prepare Week 2 plan",
     "4", "0", "REPORT TO ALKESH: Friday update #1 — share audit results, competitor data, prospect list."),
    ("", "", "WEEKEND", "", "REST", "No outreach on weekends", "0", "0", "Review responses if any come in. Don't reply until Monday."),

    # Week 2 header
    ("", "", "WEEK 2: OUTREACH BEGINS", "", "", "", "", "", ""),
    ("6", "Day 6", "Monday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + HARO",
     "1. Check for responses from Friday emails\n2. Sign up for HARO (connectively.us)\n3. Send 4 NEW outreach emails (guest post pitches)\n4. Respond to 2 HARO queries",
     "4", "0", "Start HARO early — responses take time but yield high-DR links."),
    ("7", "Day 7", "Tuesday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + FOLLOW-UP",
     "1. Send 4 NEW outreach emails\n2. Send Follow-up #1 to Day 5 emails (3 days later)\n3. Respond to any HARO queries\n4. Research 5 new prospects",
     "4", "4", "REPORT TO ALKESH: Tuesday update #2 — outreach started, X emails sent."),
    ("8", "Day 8", "Wednesday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + CONTENT",
     "1. Send 4 NEW outreach emails\n2. If anyone said YES — start writing article\n3. Respond to HARO queries\n4. Research 5 new prospects",
     "4", "0", "If you get a YES, prioritize writing content immediately."),
    ("9", "Day 9", "Thursday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + FOLLOW-UP",
     "1. Send 4 NEW outreach emails\n2. Send Follow-up #1 to Monday emails\n3. Continue writing content for any YES responses\n4. Respond to HARO",
     "4", "4", "Follow-up emails get 30% more responses than first emails."),
    ("10", "Day 10", "Friday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + REPORT",
     "1. Send 4 NEW outreach emails\n2. Send Follow-up #1 to Tuesday emails\n3. Submit any completed articles\n4. Prepare weekly report for Alkesh",
     "4", "4", "REPORT TO ALKESH: Friday update #2 — emails sent, responses received, content status."),
    ("", "", "WEEKEND", "", "REST", "Review responses. Plan next week.", "0", "0", ""),

    # Week 3 header
    ("", "", "WEEK 3: SCALE OUTREACH", "", "", "", "", "", ""),
    ("11", "Day 11", "Monday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + FOLLOW-UP",
     "1. Send 4 NEW outreach emails\n2. Send Follow-up #2 (final) to Day 5 emails\n3. Send Follow-up #1 to Wednesday emails\n4. HARO responses",
     "4", "4-8", "Week 3 is when first links typically go live."),
    ("12", "Day 12", "Tuesday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + CONTENT",
     "1. Send 4 NEW outreach emails\n2. Write/submit articles for YES responses\n3. Follow-up on pending submissions\n4. HARO responses",
     "4", "2-4", "REPORT TO ALKESH: Tuesday update #3 — first links may be live!"),
    ("13", "Day 13", "Wednesday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + VERIFY",
     "1. Send 4 NEW outreach emails\n2. Verify any published links (check dofollow)\n3. Document acquired links in backlink log\n4. Follow-ups",
     "4", "2-4", "Every published link must be verified and documented immediately."),
    ("14", "Day 14", "Thursday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + FOLLOW-UP",
     "1. Send 4 NEW outreach emails\n2. Send follow-ups to previous week's emails\n3. Continue content creation for YES responses\n4. HARO",
     "4", "4-6", "By now you should have 1-2 links acquired or in pipeline."),
    ("15", "Day 15", "Friday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + REPORT",
     "1. Send 4 NEW outreach emails\n2. Final follow-ups on old prospects\n3. Update backlink log\n4. Weekly report to Alkesh",
     "4", "2-4", "REPORT TO ALKESH: Friday update #3 — links acquired, pipeline status."),
    ("", "", "WEEKEND", "", "REST", "Review and plan final week", "0", "0", ""),

    # Week 4 header
    ("", "", "WEEK 4: CLOSE & DELIVER", "", "", "", "", "", ""),
    ("16", "Day 16", "Monday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + CLOSE",
     "1. Send 4 NEW outreach emails\n2. Follow up on all pending YES responses\n3. Push for article publication on accepted posts\n4. HARO responses",
     "4", "4-6", "Push to close any pending placements before month end."),
    ("17", "Day 17", "Tuesday", "9AM-10AM\n10AM-12PM\n1PM-3PM", "OUTREACH + CONTENT",
     "1. Send 4 NEW outreach emails\n2. Submit final articles\n3. Verify all live links\n4. Update tracking spreadsheet",
     "4", "2-4", "REPORT TO ALKESH: Tuesday update #4 — final week push."),
    ("18", "Day 18", "Wednesday", "9AM-12PM\n1PM-3PM", "CONTENT + VERIFY",
     "1. Send 2 NEW outreach emails\n2. Focus on getting pending articles published\n3. Verify all links\n4. Build Month 2 prospect list",
     "2", "2-4", "Start building Month 2 pipeline while closing Month 1."),
    ("19", "Day 19", "Thursday", "9AM-12PM\n1PM-3PM", "FINAL PUSH",
     "1. Send 2 NEW outreach emails\n2. Final follow-ups on everything pending\n3. Verify and document all links\n4. Prepare monthly report",
     "2", "4-6", "Last push to hit 4-link target."),
    ("20", "Day 20", "Friday", "9AM-12PM\n1PM-3PM", "MONTHLY REPORT",
     "1. Finalize all link verifications\n2. Complete Month 1 backlink report\n3. Referral traffic snapshot\n4. Present Month 2 plan to Alkesh",
     "0", "0", "REPORT TO ALKESH: Full Month 1 Report — links acquired, traffic data, Month 2 strategy."),
]

row = 6
for data in week1_data:
    day, date, weekday, time, task_type, actions, emails, followups, notes = data

    # Week headers
    if "WEEK" in weekday and day == "":
        ws1.merge_cells(f'A{row}:I{row}')
        sc(ws1, row, 1, weekday, Font(name='Calibri', bold=True, size=13, color=WHITE), center, dark_fill)
        for c in range(2, 10):
            ws1.cell(row=row, column=c).fill = dark_fill
            ws1.cell(row=row, column=c).border = thin_border
        row += 1
        continue

    # Weekend rows
    if weekday == "WEEKEND":
        fill = light_gray_fill
        for c in range(1, 10):
            ws1.cell(row=row, column=c).fill = fill
            ws1.cell(row=row, column=c).border = thin_border
        sc(ws1, row, 1, "", small_font, center, fill)
        sc(ws1, row, 2, "", small_font, center, fill)
        sc(ws1, row, 3, "WEEKEND", Font(name='Calibri', bold=True, italic=True, size=11, color="999999"), center, fill)
        sc(ws1, row, 4, "", small_font, center, fill)
        sc(ws1, row, 5, "REST", Font(name='Calibri', italic=True, size=11, color="999999"), center, fill)
        sc(ws1, row, 6, actions, small_font, left_align, fill)
        sc(ws1, row, 7, "0", small_font, center, fill)
        sc(ws1, row, 8, "0", small_font, center, fill)
        sc(ws1, row, 9, notes, small_font, left_align, fill)
        row += 1
        continue

    # Regular day rows
    dfill = day_fills.get(weekday, white_fill)
    sc(ws1, row, 1, day, bold_font, center, dfill)
    sc(ws1, row, 2, date, bold_font, center, dfill)
    sc(ws1, row, 3, weekday, Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE), center, dfill)
    sc(ws1, row, 4, time, small_font, center, dfill)
    sc(ws1, row, 5, task_type, Font(name='Calibri', bold=True, size=10, color=DARK_TEXT), center, dfill)
    sc(ws1, row, 6, actions, small_font, left_top, dfill)
    email_font = big_blue if str(emails) != "0" else small_font
    sc(ws1, row, 7, emails, email_font, center, dfill)
    fu_font = Font(name='Calibri', bold=True, size=11, color=ORANGE) if str(followups) != "0" else small_font
    sc(ws1, row, 8, followups, fu_font, center, dfill)
    # Highlight report days
    if "REPORT TO ALKESH" in notes:
        sc(ws1, row, 9, notes, Font(name='Calibri', bold=True, size=10, color=RED), left_top, yellow_fill)
    else:
        sc(ws1, row, 9, notes, small_font, left_top, dfill)

    ws1.row_dimensions[row].height = 65
    row += 1

# Summary at bottom
row += 1
ws1.merge_cells(f'A{row}:I{row}')
sc(ws1, row, 1, "MONTH 1 TOTALS", Font(name='Calibri', bold=True, size=14, color=WHITE), center, dark_fill)
for c in range(2, 10):
    ws1.cell(row=row, column=c).fill = dark_fill
    ws1.cell(row=row, column=c).border = thin_border

row += 1
totals = [
    ("Total Working Days", "20 days"),
    ("Total NEW Outreach Emails", "~72-80 emails"),
    ("Total Follow-up Emails", "~40-55 follow-ups"),
    ("Total Emails Sent (all types)", "~120-135 emails"),
    ("HARO Responses", "~15-20 responses"),
    ("Expected YES Responses", "4-8 (5-10% of outreach)"),
    ("Target Links Acquired", "4 backlinks (DR 60+)"),
    ("Articles Written", "4-6 guest posts"),
    ("Reports to Alkesh", "8 reports (2x per week)"),
]

for i, (label, value) in enumerate(totals):
    r = row + i
    fill = light_green_fill if "Target" in label else (light_gray_fill if i % 2 == 0 else white_fill)
    sc(ws1, r, 1, "", normal_font, center, fill)
    ws1.merge_cells(f'B{r}:E{r}')
    sc(ws1, r, 2, label, bold_font, left_align, fill)
    for c in range(3, 6):
        ws1.cell(row=r, column=c).fill = fill
        ws1.cell(row=r, column=c).border = thin_border
    ws1.merge_cells(f'F{r}:I{r}')
    sc(ws1, r, 6, value, big_blue if "Target" not in label else big_green, center, fill)
    for c in range(7, 10):
        ws1.cell(row=r, column=c).fill = fill
        ws1.cell(row=r, column=c).border = thin_border


# =====================================================
# SHEET 2: DAILY TIME BLOCKS
# =====================================================
ws2 = wb.create_sheet("Daily Time Blocks")
ws2.sheet_properties.tabColor = BRAND_GREEN

widths2 = [15, 35, 35]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:C1')
ws2.cell(row=1, column=1, value="IDEAL DAILY SCHEDULE — OUTREACH TIME BLOCKS").font = title_font
ws2.cell(row=1, column=1).fill = light_blue_fill
ws2.cell(row=1, column=1).alignment = center

ws2.merge_cells('A2:C2')
ws2.cell(row=2, column=1, value="Best time to send outreach emails: Tuesday-Thursday, 9-11 AM recipient's time zone (EST for US)").font = Font(name='Calibri', bold=True, size=11, color=GREEN)
ws2.cell(row=2, column=1).alignment = center

style_header(ws2, 4, 1, 3)
ws2.cell(row=4, column=1, value="Time Block")
ws2.cell(row=4, column=2, value="Activity")
ws2.cell(row=4, column=3, value="Details")

schedule = [
    ("9:00 - 9:30 AM", "CHECK INBOX & RESPOND", "1. Check for replies from previous outreach\n2. Respond to YES replies immediately\n3. Handle any editor questions or revision requests\n4. Check HARO daily digest email"),
    ("9:30 - 10:30 AM", "SEND NEW OUTREACH EMAILS", "1. Send 4 personalized outreach emails\n2. Each email should be customized (not copy-paste)\n3. Reference specific article on their site\n4. Use email templates but personalize 20-30%"),
    ("10:30 - 11:00 AM", "SEND FOLLOW-UPS", "1. Follow up on emails sent 3-4 days ago (Follow-up #1)\n2. Follow up on emails sent 10 days ago (Follow-up #2 — final)\n3. Update tracking spreadsheet with status"),
    ("11:00 - 12:00 PM", "CONTENT WRITING", "1. Write guest post articles for approved placements\n2. 1,500-2,000 words per article\n3. Include natural OJCommerce backlink\n4. Professional, trust-driven, value-focused tone"),
    ("12:00 - 1:00 PM", "LUNCH BREAK", "Take a break!"),
    ("1:00 - 2:00 PM", "PROSPECT RESEARCH", "1. Find 5 new potential target websites\n2. Verify DR 60+, US traffic, furniture niche\n3. Find contact emails via Hunter.io\n4. Add to prospect spreadsheet"),
    ("2:00 - 2:30 PM", "HARO RESPONSES", "1. Check afternoon HARO digest\n2. Respond to 1-2 relevant journalist queries\n3. Provide expert quotes as OJCommerce representative"),
    ("2:30 - 3:00 PM", "TRACK & DOCUMENT", "1. Update outreach tracking spreadsheet\n2. Verify any newly published links\n3. Document acquired links in backlink log\n4. Prepare notes for Alkesh's report (if report day)"),
    ("3:00 - 3:30 PM", "REPORTING (Tue & Fri only)", "1. Tuesday: Send mid-week update to Alkesh\n2. Friday: Send end-of-week update to Alkesh\n3. Include: emails sent, responses, links acquired, pipeline"),
]

for i, (time, activity, details) in enumerate(schedule):
    row = 5 + i
    if "LUNCH" in activity:
        fill = light_gray_fill
    elif "SEND NEW" in activity:
        fill = light_blue_fill
    elif "REPORTING" in activity:
        fill = yellow_fill
    elif "CONTENT" in activity:
        fill = light_green_fill
    else:
        fill = white_fill if i % 2 == 1 else light_gray_fill
    sc(ws2, row, 1, time, bold_font, center, fill)
    sc(ws2, row, 2, activity, Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE), center, fill)
    sc(ws2, row, 3, details, small_font, left_top, fill)
    ws2.row_dimensions[row].height = 70


# =====================================================
# SHEET 3: WEEKLY OVERVIEW (ALL 4 MONTHS)
# =====================================================
ws3 = wb.create_sheet("12-Month Weekly Overview")
ws3.sheet_properties.tabColor = PURPLE

widths3 = [10, 12, 14, 14, 14, 14, 12, 12, 12, 25]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells('A1:J1')
ws3.cell(row=1, column=1, value="OJCOMMERCE — 12-MONTH WEEKLY OUTREACH OVERVIEW").font = title_font
ws3.cell(row=1, column=1).fill = light_blue_fill
ws3.cell(row=1, column=1).alignment = center

style_header(ws3, 3, 1, 10)
overview_headers = ["Month", "Week", "New Emails\n/Week", "Follow-ups\n/Week", "HARO\nResponses", "Total Emails\n/Week", "Links\nTarget", "Cumulative\nLinks", "Referral\nTraffic %", "Key Focus"]
for i, h in enumerate(overview_headers, 1):
    ws3.cell(row=3, column=i, value=h)

monthly_weeks = [
    # Month 1
    ("", "MONTH 1: FOUNDATION (Target: 4 links)", "", "", "", "", "", "", "", ""),
    ("M1", "Week 1", "4", "0", "0", "4", "0", "0", "0.28%", "Setup, audit, competitor research, prospect list"),
    ("M1", "Week 2", "20", "4", "4", "28", "0-1", "0-1", "0.28%", "Active outreach begins, HARO signup"),
    ("M1", "Week 3", "20", "12", "4", "36", "1-2", "1-3", "0.30%", "Scale outreach, first YES responses, content writing"),
    ("M1", "Week 4", "12", "10", "4", "26", "1-2", "2-4", "0.43%", "Close placements, verify links, monthly report"),

    # Month 2
    ("", "MONTH 2: SCALE (Target: 6 links)", "", "", "", "", "", "", "", ""),
    ("M2", "Week 5", "24", "8", "4", "36", "1-2", "5-6", "0.50%", "New prospect batch, continued outreach"),
    ("M2", "Week 6", "24", "12", "4", "40", "1-2", "6-8", "0.55%", "Scale up, follow-ups from Week 5"),
    ("M2", "Week 7", "24", "12", "4", "40", "1-2", "7-9", "0.60%", "Content creation, link verification"),
    ("M2", "Week 8", "20", "10", "3", "33", "1-2", "8-10", "0.73%", "Close Month 2, monthly report"),

    # Month 3
    ("", "MONTH 3: MOMENTUM (Target: 8 links)", "", "", "", "", "", "", "", ""),
    ("M3", "Week 9", "30", "10", "5", "45", "2", "10-12", "0.85%", "Increased volume, Tier 2 publications targeted"),
    ("M3", "Week 10", "30", "14", "5", "49", "2", "12-14", "0.95%", "Digital PR pitches, HARO acceleration"),
    ("M3", "Week 11", "30", "14", "5", "49", "2", "14-16", "1.10%", "Content production at scale"),
    ("M3", "Week 12", "24", "12", "4", "40", "2", "16-18", "1.23%", "Close Month 3, quarterly review"),

    # Month 4-6
    ("", "MONTHS 4-6: OPTIMIZE (Target: 10-12 links/month)", "", "", "", "", "", "", "", ""),
    ("M4-6", "Weeks 13-24", "30-40/wk", "14-18/wk", "5-6/wk", "50-60/wk", "2-3/wk", "28-50", "1.96%-3.93%", "Consistent output, higher DR targets, digital PR"),

    # Month 7-9
    ("", "MONTHS 7-9: ACCELERATE (Target: 12-14 links/month)", "", "", "", "", "", "", "", ""),
    ("M7-9", "Weeks 25-36", "35-45/wk", "16-20/wk", "6-8/wk", "55-70/wk", "3-4/wk", "62-88", "4.99%-7.20%", "5% TARGET HIT (Month 7). Premium DR 70-80+ links. Scale everything."),

    # Month 10-12
    ("", "MONTHS 10-12: MAXIMIZE (Target: 14-15 links/month)", "", "", "", "", "", "", "", ""),
    ("M10-12", "Weeks 37-48", "40-50/wk", "18-22/wk", "6-8/wk", "65-80/wk", "3-4/wk", "102-131", "8.35%-9.99%", "10% STRETCH TARGET (Month 12). Elite placements. 131 total links."),
]

row = 4
for data in monthly_weeks:
    month, week, new_em, followup, haro, total, links, cumul, ref, focus = data

    # Month headers
    if "MONTH" in week:
        ws3.merge_cells(f'A{row}:J{row}')
        if "TARGET HIT" in focus or "STRETCH" in focus:
            fill = light_green_fill
            font = Font(name='Calibri', bold=True, size=12, color=GREEN)
        else:
            fill = dark_fill
            font = Font(name='Calibri', bold=True, size=12, color=WHITE)
        sc(ws3, row, 1, week, font, center, fill)
        for c in range(2, 11):
            ws3.cell(row=row, column=c).fill = fill
            ws3.cell(row=row, column=c).border = thin_border
        row += 1
        continue

    alt = light_gray_fill if row % 2 == 0 else white_fill
    if "TARGET HIT" in focus or "STRETCH" in focus:
        alt = light_green_fill

    sc(ws3, row, 1, month, bold_font, center, alt)
    sc(ws3, row, 2, week, bold_font, center, alt)
    sc(ws3, row, 3, new_em, big_blue, center, alt)
    sc(ws3, row, 4, followup, Font(name='Calibri', bold=True, size=11, color=ORANGE), center, alt)
    sc(ws3, row, 5, haro, normal_font, center, alt)
    sc(ws3, row, 6, total, bold_font, center, alt)
    sc(ws3, row, 7, links, big_green, center, alt)
    sc(ws3, row, 8, cumul, big_blue, center, alt)
    try:
        ref_val = float(ref.replace('%','').split('-')[0])
    except (ValueError, IndexError):
        ref_val = 0
    sc(ws3, row, 9, ref, Font(name='Calibri', bold=True, size=11, color=GREEN if ref_val > 2 else BRAND_BLUE), center, alt)
    sc(ws3, row, 10, focus, small_font, left_align, alt)
    ws3.row_dimensions[row].height = 35
    row += 1


# =====================================================
# SHEET 4: OUTREACH TRACKER
# =====================================================
ws4 = wb.create_sheet("Outreach Tracker")
ws4.sheet_properties.tabColor = ORANGE

widths4 = [6, 14, 25, 20, 10, 12, 12, 14, 14, 14, 14, 25]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells('A1:L1')
ws4.cell(row=1, column=1, value="OUTREACH EMAIL TRACKER — LOG EVERY EMAIL HERE").font = title_font
ws4.cell(row=1, column=1).fill = light_blue_fill
ws4.cell(row=1, column=1).alignment = center

ws4.merge_cells('A2:L2')
ws4.cell(row=2, column=1, value="Track every outreach email, follow-up, response, and result. Update daily.").font = small_font
ws4.cell(row=2, column=1).alignment = center

style_header(ws4, 4, 1, 12)
tracker_headers = ["#", "Date Sent", "Target Website", "Contact Person\n& Email", "DR", "Email Type", "Status", "Follow-up #1\nDate", "Follow-up #2\nDate", "Response\nDate", "Result", "Notes"]
for i, h in enumerate(tracker_headers, 1):
    ws4.cell(row=4, column=i, value=h)

# Sample entries
samples = [
    ("1", "Mar 2, 2026", "homedesignblog.com", "sarah@homedesignblog.com", "65", "Guest Post", "YES - Writing", "Mar 5", "N/A", "Mar 6", "LINK ACQUIRED", "Article published Mar 15"),
    ("2", "Mar 2, 2026", "furnituretips.com", "editor@furnituretips.com", "62", "Guest Post", "No Reply", "Mar 5", "Mar 12", "-", "MOVED ON", "No response after 2 follow-ups"),
    ("3", "Mar 2, 2026", "interiordesignmag.com", "james@interiordesignmag.com", "71", "Guest Post", "YES - Pending", "Mar 5", "N/A", "Mar 7", "IN PROGRESS", "Article submitted, waiting for publish"),
    ("4", "Mar 2, 2026", "homedecorinspo.com", "info@homedecorinspo.com", "58", "Guest Post", "REJECTED", "-", "-", "-", "BELOW DR 60", "DR too low — shouldn't have sent"),
    ("5", "", "", "", "", "", "", "", "", "", "", ""),
    ("6", "", "", "", "", "", "", "", "", "", "", ""),
    ("7", "", "", "", "", "", "", "", "", "", "", ""),
    ("8", "", "", "", "", "", "", "", "", "", "", ""),
    ("9", "", "", "", "", "", "", "", "", "", "", ""),
    ("10", "", "", "", "", "", "", "", "", "", "", ""),
]

for i, sample in enumerate(samples):
    row = 5 + i
    for j, val in enumerate(sample, 1):
        cell = ws4.cell(row=row, column=j, value=val)
        cell.border = thin_border
        cell.alignment = left_align
        cell.font = normal_font
        if i < 4:
            if val == "YES - Writing" or val == "LINK ACQUIRED":
                cell.font = Font(name='Calibri', bold=True, size=11, color=GREEN)
                cell.fill = light_green_fill
            elif val == "No Reply" or val == "MOVED ON":
                cell.font = Font(name='Calibri', size=11, color="999999")
            elif val == "YES - Pending" or val == "IN PROGRESS":
                cell.font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
                cell.fill = light_blue_fill
            elif val == "REJECTED" or val == "BELOW DR 60":
                cell.font = Font(name='Calibri', size=11, color=RED)
                cell.fill = light_red_fill
        else:
            cell.fill = yellow_fill  # Empty rows highlighted for input

# Add 90 more empty rows for tracking
for i in range(10, 100):
    row = 5 + i
    for j in range(1, 13):
        cell = ws4.cell(row=row, column=j, value="")
        cell.border = thin_border
    ws4.cell(row=row, column=1, value=str(i + 1)).font = small_font
    ws4.cell(row=row, column=1).alignment = center


# ============ SAVE ============
filepath = "/Users/shamique/Downloads/shridhar/Elevro/OJCommerce_Outreach_Schedule.xlsx"
wb.save(filepath)
print(f"File saved: {filepath}")
