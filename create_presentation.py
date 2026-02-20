import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ COLOR DEFINITIONS ============
DARK_BLUE = "0A1628"
BRAND_BLUE = "0066FF"
BRAND_GREEN = "00C9A7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F7FAFC"
MEDIUM_GRAY = "E2E8F0"
DARK_TEXT = "333333"
RED = "E53E3E"
GREEN = "38A169"
GOLD = "FFD700"
LIGHT_BLUE_BG = "E8F4FD"
LIGHT_GREEN_BG = "C6F6D5"
LIGHT_RED_BG = "FFE0E0"
PURPLE = "805AD5"

# ============ STYLE HELPERS ============
header_font = Font(name='Calibri', bold=True, size=14, color=WHITE)
title_font = Font(name='Calibri', bold=True, size=24, color=WHITE)
subtitle_font = Font(name='Calibri', size=14, color=BRAND_GREEN)
heading_font = Font(name='Calibri', bold=True, size=16, color=DARK_BLUE)
subheading_font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
normal_font = Font(name='Calibri', size=11, color=DARK_TEXT)
bold_font = Font(name='Calibri', bold=True, size=11, color=DARK_TEXT)
small_font = Font(name='Calibri', size=10, color="666666")
white_bold = Font(name='Calibri', bold=True, size=12, color=WHITE)
green_bold = Font(name='Calibri', bold=True, size=12, color=GREEN)
red_bold = Font(name='Calibri', bold=True, size=12, color=RED)
big_number = Font(name='Calibri', bold=True, size=20, color=BRAND_GREEN)
blue_number = Font(name='Calibri', bold=True, size=20, color=BRAND_BLUE)

dark_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
blue_fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type='solid')
green_fill = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type='solid')
light_blue_fill = PatternFill(start_color=LIGHT_BLUE_BG, end_color=LIGHT_BLUE_BG, fill_type='solid')
light_green_fill = PatternFill(start_color=LIGHT_GREEN_BG, end_color=LIGHT_GREEN_BG, fill_type='solid')
light_red_fill = PatternFill(start_color=LIGHT_RED_BG, end_color=LIGHT_RED_BG, fill_type='solid')
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header_row(ws, row, cols, fill=dark_fill, font=white_bold):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center_align
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        if alt:
            cell.fill = light_gray_fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================
# SLIDE 1: TITLE
# =====================================================
ws1 = wb.active
ws1.title = "1. Title"
ws1.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws1, [60])

for r in range(1, 15):
    ws1.cell(row=r, column=1).fill = dark_fill

ws1.cell(row=3, column=1, value="Prepared for").font = Font(name='Calibri', size=12, color="A0AEC0")
ws1.cell(row=3, column=1).alignment = center_align
ws1.cell(row=5, column=1, value="OJCommerce").font = Font(name='Calibri', bold=True, size=32, color=WHITE)
ws1.cell(row=5, column=1).alignment = center_align
ws1.cell(row=7, column=1, value="SEO Backlink Acquisition & Referral Traffic Growth").font = Font(name='Calibri', bold=True, size=18, color=BRAND_GREEN)
ws1.cell(row=7, column=1).alignment = center_align
ws1.cell(row=9, column=1, value="Strategic Proposal & Demo").font = Font(name='Calibri', size=14, color="A0AEC0")
ws1.cell(row=9, column=1).alignment = center_align
ws1.cell(row=10, column=1, value="Friday, February 20, 2026 | 3:30 PM").font = Font(name='Calibri', size=12, color="A0AEC0")
ws1.cell(row=10, column=1).alignment = center_align
ws1.cell(row=12, column=1, value="Presented by: Shamique").font = Font(name='Calibri', size=12, color=WHITE)
ws1.cell(row=12, column=1).alignment = center_align
ws1.cell(row=13, column=1, value="https://www.ojcommerce.com").font = Font(name='Calibri', size=11, color=BRAND_GREEN)
ws1.cell(row=13, column=1).alignment = center_align

# =====================================================
# SLIDE 2: AGENDA
# =====================================================
ws2 = wb.create_sheet("2. Agenda")
ws2.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws2, [8, 45, 40])

ws2.merge_cells('A1:C1')
ws2.cell(row=1, column=1, value="AGENDA").font = heading_font
ws2.cell(row=1, column=1).alignment = center_align
ws2.cell(row=1, column=1).fill = light_blue_fill

agenda_items = [
    ("1", "Understanding OJCommerce", "Business goals, challenges & current state"),
    ("2", "Current State Audit", "Where OJCommerce stands today — 122K users, 0.28% referral"),
    ("3", "Competitor Analysis", "Rooms To Go, Wayfair, Ashley, Bed Bath & Beyond, One Stop Bedroom"),
    ("4", "Backlink Gap & Opportunities", "Sites linking to competitors but NOT to OJCommerce"),
    ("5", "Our Strategy & Methodology", "White-hat approach, DR 60+ only, per-link model"),
    ("6", "Anchor Text & Link Strategy", "40% brand/homepage, natural distribution"),
    ("7", "90-Day Execution Roadmap", "Month-by-month action plan"),
    ("8", "Reporting & Transparency", "Twice-weekly updates, monthly reports"),
    ("9", "Per-Link Pricing Model", "DR-based pricing tiers"),
    ("10", "Q&A + Next Steps", "Discussion and kickoff plan"),
]

style_header_row(ws2, 3, 3)
ws2.cell(row=3, column=1, value="#")
ws2.cell(row=3, column=2, value="Topic")
ws2.cell(row=3, column=3, value="Key Points")

for i, (num, topic, detail) in enumerate(agenda_items):
    row = i + 4
    ws2.cell(row=row, column=1, value=num).font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
    ws2.cell(row=row, column=1).alignment = center_align
    ws2.cell(row=row, column=2, value=topic).font = bold_font
    ws2.cell(row=row, column=3, value=detail).font = normal_font
    style_data_row(ws2, row, 3, alt=(i % 2 == 0))

# =====================================================
# SLIDE 3: THE CHALLENGE
# =====================================================
ws3 = wb.create_sheet("3. The Challenge")
ws3.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws3, [45, 45])

ws3.merge_cells('A1:B1')
ws3.cell(row=1, column=1, value="UNDERSTANDING OJCOMMERCE'S CHALLENGE").font = heading_font
ws3.cell(row=1, column=1).fill = light_blue_fill

ws3.cell(row=3, column=1, value="THE PROBLEM").font = Font(name='Calibri', bold=True, size=13, color=RED)
ws3.cell(row=3, column=1).fill = light_red_fill
ws3.cell(row=3, column=2, value="THE OPPORTUNITY").font = Font(name='Calibri', bold=True, size=13, color=GREEN)
ws3.cell(row=3, column=2).fill = light_green_fill

problems = [
    "Current referral traffic is only 0.28% (~342 visitors/month)",
    "Previous agencies delivered low-quality, high-volume backlinks",
    "No measurable referral traffic impact from past efforts",
    "No existing backlinks or partnerships to build on",
    "Competitors (Wayfair, Rooms To Go) have massive link profiles",
]

opportunities = [
    "122,173 monthly organic users — strong foundation to build on",
    "Furniture eCommerce market is growing rapidly in the US",
    "5 competitors identified — rich backlink gap opportunities",
    "Clean slate — no Google penalties, no toxic links to clean up",
    "Per-link model ensures every dollar spent = quality result",
]

for i, (prob, opp) in enumerate(zip(problems, opportunities)):
    row = 4 + i
    ws3.cell(row=row, column=1, value=f"  {prob}").font = normal_font
    ws3.cell(row=row, column=1).alignment = left_align
    ws3.cell(row=row, column=2, value=f"  {opp}").font = normal_font
    ws3.cell(row=row, column=2).alignment = left_align

ws3.merge_cells('A10:B10')
ws3.cell(row=10, column=1, value="GOAL: Grow referral traffic from 0.28% to 5-10% by acquiring DR 60+ furniture-relevant backlinks from sites with real, verifiable US traffic.").font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
ws3.cell(row=10, column=1).fill = light_blue_fill
ws3.cell(row=10, column=1).alignment = left_align

# =====================================================
# SLIDE 4: CURRENT STATE AUDIT
# =====================================================
ws4 = wb.create_sheet("4. Current State")
ws4.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws4, [30, 25, 30])

ws4.merge_cells('A1:C1')
ws4.cell(row=1, column=1, value="OJCOMMERCE — CURRENT STATE AUDIT").font = heading_font
ws4.cell(row=1, column=1).fill = light_blue_fill

ws4.merge_cells('A2:C2')
ws4.cell(row=2, column=1, value="Website: https://www.ojcommerce.com").font = small_font

metrics = [
    ("Monthly Organic Traffic", "122,173 users", "Strong organic foundation"),
    ("Current Referral Traffic", "0.28% (~342 users/month)", "Critical — needs massive improvement"),
    ("Target Referral Traffic", "5-10% (6,100 - 12,200 users/month)", "17x to 35x growth required"),
    ("Existing Backlinks/Partnerships", "None reported", "Building from scratch — clean slate"),
    ("Blog Content", "Limited", "Mainly commercial & category pages"),
    ("Google Penalties", "None", "Clean record — no risk factors"),
    ("Geographic Focus", "United States (nationwide)", "US-focused outreach required"),
    ("Minimum DR Threshold", "DR 60+ only", "High quality bar — strict requirement"),
]

style_header_row(ws4, 4, 3)
ws4.cell(row=4, column=1, value="Metric")
ws4.cell(row=4, column=2, value="Current Value")
ws4.cell(row=4, column=3, value="Observation")

for i, (metric, value, obs) in enumerate(metrics):
    row = 5 + i
    ws4.cell(row=row, column=1, value=metric).font = bold_font
    ws4.cell(row=row, column=2, value=value).font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
    ws4.cell(row=row, column=2).alignment = center_align
    ws4.cell(row=row, column=3, value=obs).font = normal_font
    style_data_row(ws4, row, 3, alt=(i % 2 == 0))

ws4.merge_cells('A14:C14')
ws4.cell(row=14, column=1, value="KEY FINDING: OJCommerce has strong organic traffic but near-zero referral traffic. The backlink profile is essentially non-existent, presenting a massive growth opportunity.").font = Font(name='Calibri', bold=True, size=11, color=RED)
ws4.cell(row=14, column=1).fill = light_red_fill
ws4.cell(row=14, column=1).alignment = left_align

# =====================================================
# SLIDE 5: COMPETITOR ANALYSIS
# =====================================================
ws5 = wb.create_sheet("5. Competitor Analysis")
ws5.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws5, [22, 15, 18, 18, 15, 20])

ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="COMPETITOR BACKLINK COMPARISON").font = heading_font
ws5.cell(row=1, column=1).fill = light_blue_fill

style_header_row(ws5, 3, 6)
headers = ["Competitor", "Domain Rating", "Total Backlinks", "Referring Domains", "Dofollow %", "Monthly Traffic"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=3, column=i, value=h)

competitors = [
    ("OJCommerce", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "122,173"),
    ("Rooms To Go", "66", "158K", "6.6K", "79%", "Check Ahrefs"),
    ("Wayfair", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs"),
    ("Ashley Furniture", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs"),
    ("Bed Bath & Beyond", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs"),
    ("One Stop Bedroom", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs", "Check Ahrefs"),
]

for i, comp in enumerate(competitors):
    row = 4 + i
    for j, val in enumerate(comp, 1):
        cell = ws5.cell(row=row, column=j, value=val)
        cell.alignment = center_align
        cell.border = thin_border
        if i == 0:  # OJCommerce row
            cell.fill = light_blue_fill
            cell.font = bold_font
        elif val == "Check Ahrefs":
            cell.font = Font(name='Calibri', italic=True, size=10, color="999999")
        else:
            cell.font = normal_font
        if i % 2 == 1 and i != 0:
            cell.fill = light_gray_fill
    ws5.cell(row=row, column=1).font = bold_font

ws5.merge_cells('A11:F11')
ws5.cell(row=11, column=1, value="INSIGHT: Rooms To Go alone has 158K backlinks from 6.6K domains (DR 66). They get links from Wikipedia (DR 97), Synchrony (DR 87), and StarWars.com (DR 83). OJCommerce needs to close this gap.").font = Font(name='Calibri', bold=True, size=10, color=BRAND_BLUE)
ws5.cell(row=11, column=1).fill = light_blue_fill
ws5.cell(row=11, column=1).alignment = left_align

ws5.merge_cells('A13:F13')
ws5.cell(row=13, column=1, value="ACTION: Run all competitor URLs through Ahrefs Free Backlink Checker and fill in 'Check Ahrefs' cells with real data before Friday's demo.").font = Font(name='Calibri', bold=True, size=10, color=RED)
ws5.cell(row=13, column=1).fill = light_red_fill
ws5.cell(row=13, column=1).alignment = left_align

# =====================================================
# SLIDE 6: BACKLINK GAP
# =====================================================
ws6 = wb.create_sheet("6. Backlink Gap")
ws6.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws6, [25, 10, 18, 22, 20])

ws6.merge_cells('A1:E1')
ws6.cell(row=1, column=1, value="BACKLINK GAP — UNTAPPED OPPORTUNITIES").font = heading_font
ws6.cell(row=1, column=1).fill = light_blue_fill

ws6.merge_cells('A2:E2')
ws6.cell(row=2, column=1, value="Sites linking to competitors but NOT to OJCommerce = immediate outreach targets").font = small_font

style_header_row(ws6, 4, 5)
gap_headers = ["Referring Domain", "DR", "Monthly Traffic", "Links to Which Competitor?", "Links to OJCommerce?"]
for i, h in enumerate(gap_headers, 1):
    ws6.cell(row=4, column=i, value=h)

gap_data = [
    ("synchrony.com", "87", "2.5M+", "Rooms To Go", "NO"),
    ("wikipedia.org", "97", "8B+", "Rooms To Go", "NO"),
    ("concoracredit.com", "53", "Check", "Rooms To Go", "NO"),
    ("[Find via Ahrefs Gap Tool]", "-", "-", "Multiple competitors", "NO"),
    ("[Find via Ahrefs Gap Tool]", "-", "-", "Multiple competitors", "NO"),
    ("[Find via Ahrefs Gap Tool]", "-", "-", "Multiple competitors", "NO"),
]

for i, (domain, dr, traffic, comp, ojc) in enumerate(gap_data):
    row = 5 + i
    ws6.cell(row=row, column=1, value=domain).font = bold_font
    ws6.cell(row=row, column=2, value=dr).font = normal_font
    ws6.cell(row=row, column=2).alignment = center_align
    ws6.cell(row=row, column=3, value=traffic).font = normal_font
    ws6.cell(row=row, column=3).alignment = center_align
    ws6.cell(row=row, column=4, value=comp).font = Font(name='Calibri', size=11, color=GREEN)
    ws6.cell(row=row, column=4).alignment = center_align
    ws6.cell(row=row, column=5, value=ojc).font = Font(name='Calibri', bold=True, size=11, color=RED)
    ws6.cell(row=row, column=5).alignment = center_align
    ws6.cell(row=row, column=5).fill = light_red_fill
    style_data_row(ws6, row, 4, alt=(i % 2 == 0))

ws6.merge_cells('A12:E12')
ws6.cell(row=12, column=1, value="OPPORTUNITY: These sites already link to furniture brands. They are high-probability targets for OJCommerce outreach. Use Ahrefs 'Link Intersect' tool to find more.").font = Font(name='Calibri', bold=True, size=10, color=GREEN)
ws6.cell(row=12, column=1).fill = light_green_fill
ws6.cell(row=12, column=1).alignment = left_align

# =====================================================
# SLIDE 7: STRATEGY & METHODOLOGY
# =====================================================
ws7 = wb.create_sheet("7. Strategy")
ws7.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws7, [8, 25, 50])

ws7.merge_cells('A1:C1')
ws7.cell(row=1, column=1, value="OUR STRATEGY & METHODOLOGY").font = heading_font
ws7.cell(row=1, column=1).fill = light_blue_fill

ws7.merge_cells('A2:C2')
ws7.cell(row=2, column=1, value="100% White-Hat  |  Google Compliant  |  DR 60+ Only  |  Quality-First").font = Font(name='Calibri', bold=True, size=11, color=BRAND_GREEN)
ws7.cell(row=2, column=1).alignment = center_align

style_header_row(ws7, 4, 3)
ws7.cell(row=4, column=1, value="Step")
ws7.cell(row=4, column=2, value="Action")
ws7.cell(row=4, column=3, value="Details")

steps = [
    ("1", "Research & Prospecting", "Identify 100+ high-authority (DR 60+), furniture-relevant domains in the US using Ahrefs, SEMrush, and manual research"),
    ("2", "Quality Vetting", "Filter every prospect: DR 60+, real US traffic (1,000+ daily users), furniture/home niche relevance, no political/adult/gambling"),
    ("3", "Client Pre-Approval", "Present vetted prospects to OJCommerce for approval BEFORE any outreach — you control what gets pursued"),
    ("4", "Personalized Outreach", "Craft custom pitches for each target — guest posts, editorial features, resource page placements, digital PR"),
    ("5", "Content Creation", "Write high-quality, furniture-relevant content with clean, professional, trust-driven tone matching OJCommerce brand"),
    ("6", "Link Placement & QA", "Ensure natural anchor text (OJCommerce/OJ Commerce), proper dofollow attribution, 40% homepage + 60% category/product pages"),
    ("7", "Track & Report", "Monitor referral traffic via GA4 reports, track every link's performance, deliver twice-weekly updates"),
    ("8", "Optimize & Scale", "Double down on high-performing channels, adjust strategy based on data, scale outreach to new prospects"),
]

for i, (num, action, detail) in enumerate(steps):
    row = 5 + i
    ws7.cell(row=row, column=1, value=num).font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
    ws7.cell(row=row, column=1).alignment = center_align
    ws7.cell(row=row, column=2, value=action).font = bold_font
    ws7.cell(row=row, column=3, value=detail).font = normal_font
    ws7.cell(row=row, column=3).alignment = left_align
    style_data_row(ws7, row, 3, alt=(i % 2 == 0))

ws7.merge_cells('A14:C14')
ws7.cell(row=14, column=1, value="LINK TYPES: Guest Posts  |  Editorial Mentions  |  Resource Pages  |  Digital PR / HARO  |  Niche Directories (DR 60+ only)").font = Font(name='Calibri', bold=True, size=10, color=BRAND_BLUE)
ws7.cell(row=14, column=1).fill = light_blue_fill

# =====================================================
# SLIDE 8: ANCHOR TEXT STRATEGY
# =====================================================
ws8 = wb.create_sheet("8. Anchor Text")
ws8.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws8, [25, 15, 30, 5, 28, 15])

ws8.merge_cells('A1:F1')
ws8.cell(row=1, column=1, value="ANCHOR TEXT & LINK PLACEMENT STRATEGY").font = heading_font
ws8.cell(row=1, column=1).fill = light_blue_fill

# Anchor text table
ws8.merge_cells('A3:C3')
ws8.cell(row=3, column=1, value="Anchor Text Distribution").font = subheading_font

style_header_row(ws8, 4, 3)
ws8.cell(row=4, column=1, value="Anchor Type")
ws8.cell(row=4, column=2, value="Target %")
ws8.cell(row=4, column=3, value="Example")

anchors = [
    ("Branded", "35-40%", '"OJCommerce", "OJ Commerce"'),
    ("Branded + Keyword", "15-20%", '"OJCommerce furniture", "OJ Commerce bedroom sets"'),
    ("Exact Match Keyword", "5-10%", '"buy dining table online", "furniture store USA"'),
    ("Partial Match", "10-15%", '"affordable bedroom furniture", "quality home furniture"'),
    ("Generic", "5-10%", '"click here", "learn more", "visit website"'),
    ("Naked URL", "10-15%", '"ojcommerce.com", "www.ojcommerce.com"'),
]

for i, (atype, pct, example) in enumerate(anchors):
    row = 5 + i
    ws8.cell(row=row, column=1, value=atype).font = bold_font
    ws8.cell(row=row, column=2, value=pct).font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
    ws8.cell(row=row, column=2).alignment = center_align
    ws8.cell(row=row, column=3, value=example).font = normal_font
    style_data_row(ws8, row, 3, alt=(i % 2 == 0))

# Link target table
ws8.merge_cells('E3:F3')
ws8.cell(row=3, column=5, value="Link Target Distribution").font = subheading_font

style_header_row(ws8, 4, 6)
ws8.cell(row=4, column=5, value="Target Page")
ws8.cell(row=4, column=6, value="Link Share")

targets = [
    ("Homepage (brand authority)", "40%  [Client specified]"),
    ("Category Pages (bedroom, living, dining)", "30-35%"),
    ("Blog / Content Pages", "15-20%"),
    ("Top Product Pages", "5-10%"),
]

for i, (page, share) in enumerate(targets):
    row = 5 + i
    ws8.cell(row=row, column=5, value=page).font = bold_font
    ws8.cell(row=row, column=5).border = thin_border
    ws8.cell(row=row, column=6, value=share).font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
    ws8.cell(row=row, column=6).alignment = center_align
    ws8.cell(row=row, column=6).border = thin_border

ws8.merge_cells('A12:F12')
ws8.cell(row=12, column=1, value="NOTE: Alkesh specified 40% of links should target brand + homepage authority. Natural, diversified anchor text ensures Google sees links as editorially earned.").font = Font(name='Calibri', bold=True, size=10, color=BRAND_BLUE)
ws8.cell(row=12, column=1).fill = light_blue_fill

# =====================================================
# SLIDE 9: 90-DAY ROADMAP
# =====================================================
ws9 = wb.create_sheet("9. 90-Day Roadmap")
ws9.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws9, [8, 25, 50, 25])

ws9.merge_cells('A1:D1')
ws9.cell(row=1, column=1, value="90-DAY EXECUTION ROADMAP").font = heading_font
ws9.cell(row=1, column=1).fill = light_blue_fill

ws9.merge_cells('A2:D2')
ws9.cell(row=2, column=1, value="Start Date: Immediate (upon approval)").font = Font(name='Calibri', bold=True, size=11, color=GREEN)

style_header_row(ws9, 4, 4)
ws9.cell(row=4, column=1, value="Week")
ws9.cell(row=4, column=2, value="Phase")
ws9.cell(row=4, column=3, value="Activities")
ws9.cell(row=4, column=4, value="Deliverables")

roadmap = [
    ("1", "MONTH 1: FOUNDATION", "", ""),
    ("1-2", "Audit & Research", "Complete backlink & site audit for ojcommerce.com\nCompetitor gap analysis (all 5 competitors)\nBuild initial prospect list (100+ DR 60+ targets)", "Audit Report\nStrategy Document"),
    ("2-3", "Setup & Initial Outreach", "Set up tracking (GA4 referral, GSC)\nBegin personalized outreach (30+ emails/week)\nPresent first prospect batch for client approval", "Prospect List\nFirst Weekly Report"),
    ("3-4", "First Placements", "Secure first 3-5 approved placements\nCreate guest post content (professional, trust-driven tone)\nFirst links go live", "First Backlink Report\n3-5 Live Links"),
    ("", "MONTH 2: SCALE", "", ""),
    ("5-6", "Scale Outreach", "Increase outreach to 50+ emails/week\nPublish 3-5 guest posts on DR 60+ sites\nLaunch digital PR / HARO initiatives", "Weekly Reports (x2/week)\n5-8 New Backlinks"),
    ("7-8", "Optimize & Expand", "Analyze which link sources drive most referral traffic\nExpand to new prospect categories\nFirst measurable referral traffic data", "Monthly Report\nReferral Traffic Data"),
    ("", "MONTH 3: OPTIMIZE", "", ""),
    ("9-10", "Double Down", "Focus on highest-performing channels\nSecure 8-12 quality backlinks\nKeyword ranking improvements visible", "Weekly Reports (x2/week)\n8-12 New Backlinks"),
    ("11-12", "Review & Plan Q2", "Full performance review vs targets\nReferral traffic trend analysis\nQuarter 2 strategy planning", "Full 90-Day Report\nQ2 Strategy Plan"),
]

for i, (week, phase, activities, deliverables) in enumerate(roadmap):
    row = 5 + i
    if phase.startswith("MONTH"):
        ws9.cell(row=row, column=1, value="").fill = dark_fill
        ws9.cell(row=row, column=2, value=phase).font = white_bold
        ws9.cell(row=row, column=2).fill = dark_fill
        ws9.cell(row=row, column=3, value="").fill = dark_fill
        ws9.cell(row=row, column=4, value="").fill = dark_fill
    else:
        ws9.cell(row=row, column=1, value=week).font = Font(name='Calibri', bold=True, size=11, color=BRAND_BLUE)
        ws9.cell(row=row, column=1).alignment = center_align
        ws9.cell(row=row, column=2, value=phase).font = bold_font
        ws9.cell(row=row, column=3, value=activities).font = normal_font
        ws9.cell(row=row, column=3).alignment = left_top
        ws9.cell(row=row, column=4, value=deliverables).font = normal_font
        ws9.cell(row=row, column=4).alignment = left_top
        style_data_row(ws9, row, 4)

ws9.merge_cells('A16:D16')
ws9.cell(row=16, column=1, value="EXPECTED 90-DAY OUTCOMES: 20-30 DR 60+ backlinks  |  DR improvement of 3-5 points  |  Referral traffic growth from 0.28% toward 2-3%  |  Keyword ranking movement").font = Font(name='Calibri', bold=True, size=10, color=GREEN)
ws9.cell(row=16, column=1).fill = light_green_fill

# =====================================================
# SLIDE 10: REPORTING
# =====================================================
ws10 = wb.create_sheet("10. Reporting")
ws10.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws10, [20, 35, 35])

ws10.merge_cells('A1:C1')
ws10.cell(row=1, column=1, value="REPORTING & TRANSPARENCY").font = heading_font
ws10.cell(row=1, column=1).fill = light_blue_fill

ws10.merge_cells('A2:C2')
ws10.cell(row=2, column=1, value="Alkesh requested: Twice-weekly updates").font = Font(name='Calibri', bold=True, size=11, color=BRAND_GREEN)

style_header_row(ws10, 4, 3)
ws10.cell(row=4, column=1, value="Frequency")
ws10.cell(row=4, column=2, value="What We Report")
ws10.cell(row=4, column=3, value="Format")

reporting = [
    ("Tuesday Update", "Outreach emails sent, responses received, links in pipeline, any blockers", "Email / Slack summary"),
    ("Friday Update", "Links acquired this week, referral traffic snapshot, DR/quality metrics, next week plan", "Email / Slack summary"),
    ("Monthly Report", "All backlinks acquired (URL, anchor, DR, traffic, type), referral traffic from GA4, keyword ranking changes, competitor comparison", "Detailed spreadsheet + summary doc"),
    ("Monthly Call", "Live walkthrough of results, strategy discussion, approval of next month's prospects", "30-min video call"),
    ("Quarterly Review", "Full performance vs targets, ROI assessment, 6-month progress toward 5-10% referral", "Comprehensive presentation"),
]

for i, (freq, what, fmt) in enumerate(reporting):
    row = 5 + i
    ws10.cell(row=row, column=1, value=freq).font = bold_font
    ws10.cell(row=row, column=2, value=what).font = normal_font
    ws10.cell(row=row, column=2).alignment = left_align
    ws10.cell(row=row, column=3, value=fmt).font = normal_font
    style_data_row(ws10, row, 3, alt=(i % 2 == 0))

ws10.merge_cells('A11:C11')
ws10.cell(row=11, column=1, value="BACKLINK LOG FORMAT (for each link acquired):").font = subheading_font

style_header_row(ws10, 12, 3)
log_headers = [("Field", "Description", "Example")]
ws10.cell(row=12, column=1, value="Field")
ws10.cell(row=12, column=2, value="Description")
ws10.cell(row=12, column=3, value="Example")

log_fields = [
    ("Placement URL", "The exact page where the link lives", "homedesignblog.com/furniture-guide"),
    ("Anchor Text", "The clickable text used", "OJCommerce"),
    ("Target URL", "Which OJCommerce page the link points to", "ojcommerce.com/bedroom-furniture"),
    ("Domain Rating (DR)", "Authority of the linking site", "DR 65"),
    ("Monthly Traffic", "Traffic of the linking domain", "85,000/month"),
    ("Link Type", "Dofollow/nofollow, guest post/editorial", "Dofollow, Guest Post"),
    ("Date Acquired", "When the link went live", "March 15, 2026"),
    ("Referral Traffic", "Visitors from this specific link (GA4)", "127 users this month"),
]

for i, (field, desc, example) in enumerate(log_fields):
    row = 13 + i
    ws10.cell(row=row, column=1, value=field).font = bold_font
    ws10.cell(row=row, column=2, value=desc).font = normal_font
    ws10.cell(row=row, column=3, value=example).font = Font(name='Calibri', italic=True, size=11, color="666666")
    style_data_row(ws10, row, 3, alt=(i % 2 == 0))

# =====================================================
# SLIDE 11: COMPLIANCE
# =====================================================
ws11 = wb.create_sheet("11. Compliance")
ws11.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws11, [45, 45])

ws11.merge_cells('A1:B1')
ws11.cell(row=1, column=1, value="QUALITY ASSURANCE & COMPLIANCE").font = heading_font
ws11.cell(row=1, column=1).fill = light_blue_fill

ws11.cell(row=3, column=1, value="WHAT WE DO").font = Font(name='Calibri', bold=True, size=13, color=GREEN)
ws11.cell(row=3, column=1).fill = light_green_fill
ws11.cell(row=3, column=2, value="WHAT WE NEVER DO").font = Font(name='Calibri', bold=True, size=13, color=RED)
ws11.cell(row=3, column=2).fill = light_red_fill

dos = [
    "Editorial guest posts on real sites with real US traffic",
    "HARO & journalist outreach for expert mentions",
    "Resource page placements on DR 60+ sites",
    "Digital PR through newsworthy content",
    "Verify every prospect for traffic, relevance, and authority",
    "Natural anchor text (OJCommerce / OJ Commerce variations)",
    "Pre-approval: you approve every placement before we proceed",
    "Full compliance with Google Webmaster Guidelines",
    "Only furniture, home, interior design, lifestyle niches",
]

donts = [
    "Private Blog Networks (PBNs)",
    "Link farms or bulk link purchases",
    "Spammy directories or low-quality submissions",
    "Comment spam or forum spam",
    "Hidden or cloaked links",
    "Political, adult, gambling, or unrelated niche sites",
    "Any site below DR 60",
    "Any technique that risks a Google penalty",
    "Volume-focused approach (Alkesh's past frustration)",
]

for i in range(max(len(dos), len(donts))):
    row = 4 + i
    if i < len(dos):
        ws11.cell(row=row, column=1, value=f"  {dos[i]}").font = normal_font
    if i < len(donts):
        ws11.cell(row=row, column=2, value=f"  {donts[i]}").font = normal_font

ws11.merge_cells('A14:B14')
ws11.cell(row=14, column=1, value="OUR GUARANTEE: Every backlink is from a real DR 60+ website, with real traffic, editorially placed, and fully Google-compliant. You approve every placement.").font = Font(name='Calibri', bold=True, size=11, color=GREEN)
ws11.cell(row=14, column=1).fill = light_green_fill

# =====================================================
# SLIDE 12: PRICING
# =====================================================
ws12 = wb.create_sheet("12. Pricing")
ws12.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws12, [25, 20, 20, 25])

ws12.merge_cells('A1:D1')
ws12.cell(row=1, column=1, value="PER-LINK PRICING MODEL").font = heading_font
ws12.cell(row=1, column=1).fill = light_blue_fill

ws12.merge_cells('A2:D2')
ws12.cell(row=2, column=1, value="As requested by Alkesh: No monthly retainer. Pay per approved, quality backlink.").font = Font(name='Calibri', bold=True, size=11, color=BRAND_GREEN)

style_header_row(ws12, 4, 4)
ws12.cell(row=4, column=1, value="DR Range")
ws12.cell(row=4, column=2, value="Price Per Link")
ws12.cell(row=4, column=3, value="What's Included")
ws12.cell(row=4, column=4, value="Typical Source")

pricing = [
    ("DR 60-70", "$[Your Price]", "Prospect research, outreach, content creation, placement, tracking", "Niche blogs, industry sites, resource pages"),
    ("DR 70-80", "$[Your Price]", "All above + premium content, higher authority targets", "Major industry publications, lifestyle magazines"),
    ("DR 80+", "$[Your Price]", "All above + digital PR, editorial features on elite sites", "Major news sites, Wikipedia-level domains"),
]

for i, (dr, price, included, source) in enumerate(pricing):
    row = 5 + i
    ws12.cell(row=row, column=1, value=dr).font = Font(name='Calibri', bold=True, size=14, color=BRAND_BLUE)
    ws12.cell(row=row, column=1).alignment = center_align
    ws12.cell(row=row, column=2, value=price).font = Font(name='Calibri', bold=True, size=14, color=RED)
    ws12.cell(row=row, column=2).alignment = center_align
    ws12.cell(row=row, column=3, value=included).font = normal_font
    ws12.cell(row=row, column=3).alignment = left_align
    ws12.cell(row=row, column=4, value=source).font = normal_font
    ws12.cell(row=row, column=4).alignment = left_align
    style_data_row(ws12, row, 4, alt=(i % 2 == 0))

ws12.merge_cells('A9:D9')
ws12.cell(row=9, column=1, value="ENGAGEMENT DETAILS").font = subheading_font

details = [
    ("Engagement Model", "Per-link pricing (as requested)"),
    ("Setup Fee", "$[Your Price] — one-time audit, strategy, prospect list development"),
    ("Minimum Commitment", "3 months recommended for measurable results"),
    ("Approval Process", "Every prospect pre-approved by OJCommerce before outreach"),
    ("Payment Terms", "Invoice upon link going live and verified"),
    ("Reporting", "Twice-weekly updates (Tuesday + Friday) + monthly report"),
    ("Start Date", "Immediate upon approval (as requested)"),
    ("Decision Maker", "Business Owner (final approval)"),
]

style_header_row(ws12, 10, 2)
ws12.cell(row=10, column=1, value="Item")
ws12.cell(row=10, column=2, value="Detail")
ws12.merge_cells('B10:D10')

for i, (item, detail) in enumerate(details):
    row = 11 + i
    ws12.cell(row=row, column=1, value=item).font = bold_font
    ws12.cell(row=row, column=1).border = thin_border
    ws12.merge_cells(f'B{row}:D{row}')
    ws12.cell(row=row, column=2, value=detail).font = normal_font
    ws12.cell(row=row, column=2).border = thin_border
    if i % 2 == 0:
        ws12.cell(row=row, column=1).fill = light_gray_fill
        ws12.cell(row=row, column=2).fill = light_gray_fill

ws12.merge_cells('A20:D20')
ws12.cell(row=20, column=1, value="NOTE: Replace $[Your Price] with your actual pricing before Friday's demo.").font = Font(name='Calibri', bold=True, size=10, color=RED)
ws12.cell(row=20, column=1).fill = light_red_fill

# =====================================================
# SLIDE 13: WHY US
# =====================================================
ws13 = wb.create_sheet("13. Why Choose Us")
ws13.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws13, [10, 35, 45])

ws13.merge_cells('A1:C1')
ws13.cell(row=1, column=1, value="WHY CHOOSE US?").font = heading_font
ws13.cell(row=1, column=1).fill = light_blue_fill

ws13.merge_cells('A3:C3')
ws13.cell(row=3, column=1, value="How We Address OJCommerce's Specific Pain Points").font = subheading_font

style_header_row(ws13, 4, 3)
ws13.cell(row=4, column=1, value="#")
ws13.cell(row=4, column=2, value="Alkesh's Pain Point")
ws13.cell(row=4, column=3, value="Our Solution")

pain_solutions = [
    ("1", "Previous agencies focused on volume over quality", "We ONLY target DR 60+ sites. Every link pre-approved by you. Quality is non-negotiable."),
    ("2", "Low-quality backlinks delivered", "Every prospect vetted for real traffic, niche relevance, and authority. No shortcuts."),
    ("3", "No measurable referral traffic impact", "We track referral traffic per link via GA4. Every link must prove its value."),
    ("4", "Lack of transparency", "Twice-weekly reports, monthly calls, full backlink log with every detail exposed."),
    ("5", "Wasted budget on junk links", "Per-link pricing = you only pay for approved, verified, quality placements."),
    ("6", "No clear strategy", "90-day roadmap with weekly milestones, clear targets, and measurable KPIs."),
]

for i, (num, pain, solution) in enumerate(pain_solutions):
    row = 5 + i
    ws13.cell(row=row, column=1, value=num).font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
    ws13.cell(row=row, column=1).alignment = center_align
    ws13.cell(row=row, column=2, value=pain).font = Font(name='Calibri', size=11, color=RED)
    ws13.cell(row=row, column=3, value=solution).font = Font(name='Calibri', size=11, color=GREEN)
    style_data_row(ws13, row, 3, alt=(i % 2 == 0))

# =====================================================
# SLIDE 14: NEXT STEPS
# =====================================================
ws14 = wb.create_sheet("14. Next Steps")
ws14.sheet_properties.tabColor = BRAND_BLUE
set_col_widths(ws14, [8, 30, 45])

ws14.merge_cells('A1:C1')
ws14.cell(row=1, column=1, value="NEXT STEPS — LET'S GET STARTED").font = heading_font
ws14.cell(row=1, column=1).fill = light_blue_fill

style_header_row(ws14, 3, 3)
ws14.cell(row=3, column=1, value="Step")
ws14.cell(row=3, column=2, value="Action")
ws14.cell(row=3, column=3, value="Timeline")

next_steps = [
    ("1", "Today's demo — align on strategy & approach", "Today (Friday)"),
    ("2", "Business Owner approval", "This week"),
    ("3", "Sign engagement agreement", "Upon approval"),
    ("4", "Kick off — audit & strategy development", "Week 1 after approval"),
    ("5", "Share priority product/category pages for targeting", "Week 1"),
    ("6", "Share GA4 & GSC reports (as Alkesh agreed)", "Week 1"),
    ("7", "First prospect list delivered for approval", "Week 2"),
    ("8", "Outreach begins", "Week 2-3"),
    ("9", "First backlinks acquired & reported", "Week 3-4"),
    ("10", "First twice-weekly report delivered", "Week 2 onwards"),
]

for i, (num, action, timeline) in enumerate(next_steps):
    row = 4 + i
    ws14.cell(row=row, column=1, value=num).font = Font(name='Calibri', bold=True, size=12, color=BRAND_BLUE)
    ws14.cell(row=row, column=1).alignment = center_align
    ws14.cell(row=row, column=2, value=action).font = bold_font
    ws14.cell(row=row, column=3, value=timeline).font = normal_font
    style_data_row(ws14, row, 3, alt=(i % 2 == 0))

ws14.merge_cells('A15:C15')
ws14.cell(row=15, column=1).fill = light_green_fill
ws14.cell(row=15, column=1, value="READY TO START IMMEDIATELY — As Alkesh requested, we can begin the audit and strategy phase within days of approval.").font = Font(name='Calibri', bold=True, size=11, color=GREEN)

ws14.merge_cells('A17:C17')
ws14.cell(row=17, column=1, value="Thank you, Alkesh. We look forward to partnering with OJCommerce.").font = Font(name='Calibri', bold=True, size=14, color=DARK_BLUE)
ws14.cell(row=17, column=1).alignment = center_align

ws14.merge_cells('A18:C18')
ws14.cell(row=18, column=1, value="Presented by: Shamique  |  [Your Agency Name]  |  [Email]  |  [Phone]").font = Font(name='Calibri', size=12, color=BRAND_BLUE)
ws14.cell(row=18, column=1).alignment = center_align

# ============ SAVE ============
filepath = "/Users/shamique/Downloads/shridhar/Elevro/OJCommerce_SEO_Demo_Presentation.xlsx"
wb.save(filepath)
print(f"File saved: {filepath}")
